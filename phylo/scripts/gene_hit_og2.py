#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gene_hit_og.py  (OrthoFinder Orthogroups.tsv as the ONLY mapping source)

皇上铁律（已落实）：
1) 仅使用 OrthoFinder 的 Orthogroups/Orthogroups.tsv
2) Orthogroups.tsv 中的成员 ID：若包含'|'，取最后一个'|'之后作为真实 protein_id
3) 禁止任何 '.' 后缀删除/推断最长转录本/isoform/base_id/prefix 模糊匹配：只做精确匹配
4) -p 模式：未命中 OG 的行统一放到输出末尾（OG=NA），与原脚本一致
5) 屏幕输出不刷屏：多 OG 冲突不逐行吐，只做汇总并写日志文件

模式A（默认，位置参数 QUERY）：
- QUERY 可以是 protein_id 字符串，也可以是“包含 protein_id 的文件”
- 输出：
  1) OUT_BASE/OG_hits_summary.tsv：query_protein_id \t OG \t Species（无表头）
  2) OUT_BASE/OG_species_counts.tsv：行=OG，列=Species，值=该 OG 在该物种下 protein_id 数量
  3) OUT_BASE/<OG>/<Species>.txt：每行一个 protein_id（无表头）

模式B（-p FILE）：
- 输入一个 TSV 文件（保留所有原始列与值），仅新增/覆盖 OG 列并按 OG 聚类排序
- 未命中行 OG=NA，统一放到输出末尾
- 输出：
  - TSV：OUT_BASE/preprocessed/<输入文件名>.ogcluster.tsv
  - XLSX：OUT_BASE/preprocessed/<输入文件名>.ogcluster.xlsx（可开关）
"""

from __future__ import annotations

import sys
import argparse
from pathlib import Path
from collections import defaultdict
from datetime import datetime


# =============================================================================
# 顶部参数区（只改这里就行；全部为“相对脚本所在目录”的相对路径）
# =============================================================================

# 文件表头开关：
# True：凡是“输入是文件”（QUERY 文件、-p 文件），默认第一行是表头
# False：文件无表头（第一行当数据）
FILE_INPUT_HAS_HEADER = True

# OrthoFinder OG 映射表（唯一映射源）
ORTHO_GROUPS_TSV_REL = "../results/orthofinder/Results_Jan21/Orthogroups/Orthogroups.tsv"

OUT_BASE_REL = "../results/out_og_hits"

HITS_SUMMARY_NAME = "OG_hits_summary.tsv"
COUNTS_MATRIX_NAME = "OG_species_counts.tsv"

# -p 模式 XLSX 输出开关
WRITE_XLSX_FOR_P = True

# 多 OG 冲突：最多记录多少个样例到日志（避免日志无限膨胀）
MAX_MULTI_OG_EXAMPLES_IN_LOG = 200

# 日志文件名（写入 OUT_BASE/logs/）
LOG_NAME = "gene_hit_og.log"

# 皇上御用20色（HEX）
ROYAL_20_HEX = [
    "#F79D93", "#95C8F2", "#F6CD96", "#9FD5CB", "#F6C6E7",
    "#A99BEF", "#F5B07E", "#A7D9F7", "#F3A6C9", "#B6E0B6",
    "#FBE3A8", "#8FE3F2", "#F4B1A8", "#8FB1F2", "#F0D07A",
    "#7FD3C8", "#F7B6D2", "#C3A3F5", "#D6E7B5", "#9E90E6",
]

# “清新化”系数：把颜色与白色混合（0~1）
FRESHEN_FACTOR = 0.45

# NA 的背景色（清淡灰）
NA_FILL_HEX = "#F2F2F2"

# =============================================================================


def sanitize_filename(name: str) -> str:
    bad = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    for ch in bad:
        name = name.replace(ch, "_")
    name = name.strip().replace(" ", "_")
    return name


def clean_cell(x: str) -> str:
    return x.strip().strip('"').strip("'")


def split_query_tokens(raw: str) -> list[str]:
    raw = raw.strip()
    if not raw:
        return []
    tmp = raw.replace(",", " ")
    toks = [t.strip() for t in tmp.split() if t.strip()]
    return toks


def resolve_existing_file(arg: str, script_dir: Path) -> Path | None:
    p_cwd = Path(arg)
    if p_cwd.exists() and p_cwd.is_file():
        return p_cwd
    p_script = script_dir / arg
    if p_script.exists() and p_script.is_file():
        return p_script
    return None


# -------------------------------------------------------------------------
# 日志：屏幕流式 + 文件落盘
# -------------------------------------------------------------------------
class Logger:
    def __init__(self, log_path: Path):
        self.log_path = log_path
        self.log_path.parent.mkdir(parents=True, exist_ok=True)
        self._fh = self.log_path.open("a", encoding="utf-8")

    def close(self) -> None:
        try:
            self._fh.close()
        except Exception:
            pass

    def _stamp(self) -> str:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def info(self, msg: str) -> None:
        line = f"[{self._stamp()}] [INFO] {msg}"
        print(line, file=sys.stderr)
        self._fh.write(line + "\n")
        self._fh.flush()

    def warn(self, msg: str) -> None:
        line = f"[{self._stamp()}] [WARN] {msg}"
        print(line, file=sys.stderr)
        self._fh.write(line + "\n")
        self._fh.flush()

    def ok(self, msg: str) -> None:
        line = f"[{self._stamp()}] [OK] {msg}"
        print(line, file=sys.stderr)
        self._fh.write(line + "\n")
        self._fh.flush()

    def error(self, msg: str) -> None:
        line = f"[{self._stamp()}] [ERROR] {msg}"
        print(line, file=sys.stderr)
        self._fh.write(line + "\n")
        self._fh.flush()

    def raw_to_log(self, msg: str) -> None:
        self._fh.write(msg + "\n")
        self._fh.flush()


# -------------------------------------------------------------------------
# ID 规范化（皇上指定：取最后一个 '|' 之后；严禁动 '.'）
# -------------------------------------------------------------------------
def normalize_pid_keep_dot(raw_id: str) -> str:
    """
    统一用于：
    - Orthogroups.tsv 单元格成员 ID
    - QUERY / -p 输入表中的 protein_id

    规则：
    1) 去掉引号/空白
    2) 若包含'|'，取最后一个'|'之后的部分
    3) 严禁对 '.' 做任何截断或推断（完整保留）
    """
    x = clean_cell(raw_id).strip()
    if not x:
        return ""
    if "|" in x:
        x = x.rsplit("|", 1)[1].strip()
    return x


# -------------------------------------------------------------------------
# 读取 OrthoFinder Orthogroups.tsv 并构建索引
# -------------------------------------------------------------------------
def read_orthogroups_header(header_line: str, log: Logger) -> list[str]:
    cols = header_line.rstrip("\n").split("\t")
    if len(cols) < 2:
        raise RuntimeError("Orthogroups.tsv 表头列数不足（至少应包含 Orthogroup + 一个物种列）。")

    first = cols[0].strip().lstrip("\ufeff").lower()
    if first not in ("orthogroup", "orthogroups", "og", "orthogroupid"):
        log.warn(f"Orthogroups.tsv 第一列列名不是常见 'Orthogroup'，实际为 '{cols[0]}'。仍将其视为 OG 列。")

    species_order = [c.strip() for c in cols[1:]]
    if any(not s for s in species_order):
        raise RuntimeError("Orthogroups.tsv 表头存在空的物种列名。")
    return species_order


def parse_member_cell(cell: str) -> list[str]:
    """
    Orthogroups.tsv 单元格成员列表解析：
    - 你示例中是 ',' 分隔，且常见 ', '。
    - 这里固定按 ',' 拆分，然后 normalize。
    """
    s = cell.strip()
    if not s:
        return []
    parts = [p.strip() for p in s.split(",")]
    out: list[str] = []
    for p in parts:
        pid = normalize_pid_keep_dot(p)
        if pid:
            out.append(pid)
    return out


def load_maps_from_orthogroups(
    og_tsv: Path,
    log: Logger,
) -> tuple[dict[str, set[str]], dict[str, dict[str, set[str]]], list[str], dict[str, list[str]]]:
    """
    返回：
    - exact_map: pid -> {OG}
    - og2sp2prot: OG -> Species -> {pid}
    - species_order: 物种列顺序
    - multi_og_examples: pid -> sorted OG list （仅记录最多 MAX_MULTI_OG_EXAMPLES_IN_LOG 个样例）
    """
    exact_map: dict[str, set[str]] = defaultdict(set)
    og2sp2prot: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))

    multi_og_examples: dict[str, list[str]] = {}
    multi_og_seen = 0

    with og_tsv.open("r", encoding="utf-8", errors="strict") as f:
        header = f.readline()
        if not header:
            raise RuntimeError("Orthogroups.tsv 是空文件。")
        species_order = read_orthogroups_header(header, log)

        for line_no, line in enumerate(f, start=2):
            line = line.rstrip("\n")
            if not line:
                continue
            parts = line.split("\t")
            if not parts:
                continue

            og = parts[0].strip()
            if not og:
                continue

            cells = parts[1:]
            if len(cells) < len(species_order):
                cells = cells + [""] * (len(species_order) - len(cells))

            for sp, cell in zip(species_order, cells):
                members = parse_member_cell(cell)
                if not members:
                    continue
                for pid in members:
                    exact_map[pid].add(og)
                    og2sp2prot[og][sp].add(pid)

    # 汇总多 OG 冲突样例（只进日志，不刷屏）
    for pid, ogs in exact_map.items():
        if len(ogs) > 1:
            multi_og_seen += 1
            if len(multi_og_examples) < MAX_MULTI_OG_EXAMPLES_IN_LOG:
                multi_og_examples[pid] = sorted(ogs)

    if multi_og_seen > 0:
        log.warn(f"Detected protein_id mapped to multiple OGs: count={multi_og_seen}. Details are written to log (up to {MAX_MULTI_OG_EXAMPLES_IN_LOG} examples).")
        if multi_og_examples:
            log.raw_to_log("---- Multi-OG examples (pid -> OGs) ----")
            for pid, ogs in list(multi_og_examples.items()):
                log.raw_to_log(f"{pid}\t{','.join(ogs)}")
            log.raw_to_log("---- End of Multi-OG examples ----")

    return exact_map, og2sp2prot, species_order, multi_og_examples


# -------------------------------------------------------------------------
# 精确匹配（只查 exact_map；多 OG 时取字典序最小，但不刷屏）
# -------------------------------------------------------------------------
def match_pid_to_single_og_exact(
    query_pid: str,
    exact_map: dict[str, set[str]],
    multi_og_counter: dict[str, int],
) -> str | None:
    q = normalize_pid_keep_dot(query_pid)
    if not q:
        return None

    ogs = exact_map.get(q, set())
    if not ogs:
        return None

    if len(ogs) == 1:
        return next(iter(ogs))

    # 多 OG：不逐条 warn，改为计数汇总；选择字典序最小 OG 保证确定性
    multi_og_counter[q] = multi_og_counter.get(q, 0) + 1
    return sorted(ogs)[0]


def resolve_query_as_ids(arg1: str, script_dir: Path) -> tuple[list[str], str]:
    fp = resolve_existing_file(arg1, script_dir)
    if fp is not None:
        ids: list[str] = []
        with fp.open("r", encoding="utf-8", errors="strict") as f:
            first = True
            for line in f:
                if first and FILE_INPUT_HAS_HEADER:
                    first = False
                    continue
                first = False
                line = line.strip()
                if not line:
                    continue
                ids.extend(split_query_tokens(line))
        return ids, "file"
    return split_query_tokens(arg1), "string"


# -------------------------------------------------------------------------
# 颜色与 XLSX 输出
# -------------------------------------------------------------------------
def hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
    h = hex_color.strip().lstrip("#")
    if len(h) != 6:
        raise RuntimeError(f"非法 HEX 颜色：{hex_color}")
    r = int(h[0:2], 16)
    g = int(h[2:4], 16)
    b = int(h[4:6], 16)
    return r, g, b


def rgb_to_hex(r: int, g: int, b: int) -> str:
    r = max(0, min(255, int(r)))
    g = max(0, min(255, int(g)))
    b = max(0, min(255, int(b)))
    return f"#{r:02X}{g:02X}{b:02X}"


def freshen_hex(hex_color: str, factor: float) -> str:
    r, g, b = hex_to_rgb(hex_color)
    r2 = r * (1 - factor) + 255 * factor
    g2 = g * (1 - factor) + 255 * factor
    b2 = b * (1 - factor) + 255 * factor
    return rgb_to_hex(r2, g2, b2)


def build_og_fill_map(ogs_in_order: list[str]) -> dict[str, str]:
    base = [freshen_hex(c, FRESHEN_FACTOR) for c in ROYAL_20_HEX]
    m: dict[str, str] = {}
    k = 0
    for og in ogs_in_order:
        if og == "NA":
            continue
        if og in m:
            continue
        m[og] = base[k % len(base)]
        k += 1
    return m


def write_xlsx_with_og_colors(xlsx_path: Path, rows: list[list[str]], has_header: bool) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "OG_cluster"

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    data_start = 2 if has_header else 1
    ogs_seen: list[str] = []
    for r_idx in range(data_start, len(rows) + 1):
        og = str(ws.cell(row=r_idx, column=1).value or "").strip()
        if not og:
            og = "NA"
        if og not in ogs_seen:
            ogs_seen.append(og)

    og2hex = build_og_fill_map(ogs_seen)

    for r_idx in range(data_start, len(rows) + 1):
        og = str(ws.cell(row=r_idx, column=1).value or "").strip()
        if not og:
            og = "NA"
        hex_color = NA_FILL_HEX if og == "NA" else og2hex.get(og, NA_FILL_HEX)
        argb = "FF" + hex_color.lstrip("#").upper()
        fill = PatternFill(patternType="solid", fgColor=argb)
        ws.cell(row=r_idx, column=1).fill = fill

    max_cols = max(len(r) for r in rows) if rows else 1
    for c in range(1, max_cols + 1):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 18 if c == 1 else 22

    wb.save(xlsx_path)


# -------------------------------------------------------------------------
# -p 模式：按 OG 聚类，NA 行集中到末尾（保留原始语义）
# -------------------------------------------------------------------------
def preprocess_file_cluster_by_og(
    in_file: Path,
    out_tsv: Path,
    out_xlsx: Path,
    exact_map: dict[str, set[str]],
    log: Logger,
) -> None:
    out_tsv.parent.mkdir(parents=True, exist_ok=True)

    rows_for_xlsx: list[list[str]] = []
    has_header = FILE_INPUT_HAS_HEADER

    multi_og_counter: dict[str, int] = {}

    with in_file.open("r", encoding="utf-8", errors="strict") as f:
        first_line = f.readline()
        if not first_line:
            raise RuntimeError(f"-p 输入文件为空：{in_file}")

        header_fields: list[str] | None = None
        og_already_first = False

        if FILE_INPUT_HAS_HEADER:
            header_fields = first_line.rstrip("\n").split("\t")
            if len(header_fields) == 0:
                raise RuntimeError(f"-p 输入文件表头为空：{in_file}")
            og_already_first = (header_fields[0].strip().lower().lstrip("\ufeff") == "og")
        else:
            data0 = first_line.rstrip("\n").split("\t")
            if len(data0) == 0:
                raise RuntimeError(f"-p 输入文件首行为空：{in_file}")
            og_already_first = (clean_cell(data0[0]).strip().lower() == "og")

        id_col = 1 if og_already_first else 0

        groups: dict[str, list[list[str]]] = {}
        og_order: list[str] = []
        unmatched_rows: list[list[str]] = []

        def add_to_group(og: str, row_fields: list[str]) -> None:
            if og not in groups:
                groups[og] = []
                og_order.append(og)
            groups[og].append(row_fields)

        def process_row(fields: list[str], line_no: int) -> None:
            if id_col >= len(fields):
                raise RuntimeError(f"-p 输入文件第 {line_no} 行列数不足：{fields}")

            pid = fields[id_col]
            og = match_pid_to_single_og_exact(pid, exact_map, multi_og_counter)
            og_val = og if og is not None else "NA"

            if og_already_first:
                new_fields = fields[:]
                new_fields[0] = og_val
            else:
                new_fields = [og_val] + fields

            if og is None:
                unmatched_rows.append(new_fields)
            else:
                add_to_group(og, new_fields)

        if FILE_INPUT_HAS_HEADER:
            header_out = header_fields if og_already_first else (["OG"] + header_fields)
            rows_for_xlsx.append(header_out)
            for line_no, line in enumerate(f, start=2):
                line = line.rstrip("\n")
                if line == "":
                    continue
                process_row(line.split("\t"), line_no)
        else:
            process_row(first_line.rstrip("\n").split("\t"), 1)
            for line_no, line in enumerate(f, start=2):
                line = line.rstrip("\n")
                if line == "":
                    continue
                process_row(line.split("\t"), line_no)

        for og in og_order:
            for row in groups[og]:
                rows_for_xlsx.append(row)
        for row in unmatched_rows:
            rows_for_xlsx.append(row)

    with out_tsv.open("w", encoding="utf-8") as w:
        for row in rows_for_xlsx:
            w.write("\t".join(row) + "\n")

    if WRITE_XLSX_FOR_P:
        write_xlsx_with_og_colors(out_xlsx, rows_for_xlsx, has_header=has_header)

    if multi_og_counter:
        total = sum(multi_og_counter.values())
        uniq = len(multi_og_counter)
        log.warn(f"-p mode: multi-OG exact-id conflicts encountered: total_hits={total}, unique_ids={uniq}. (No per-row spam; see log for sampling if needed.)")
        # 只记录少量样例到日志
        log.raw_to_log("---- -p multi-OG conflict sample (pid -> times) ----")
        for pid, times in list(sorted(multi_og_counter.items(), key=lambda x: (-x[1], x[0])))[:50]:
            log.raw_to_log(f"{pid}\t{times}")
        log.raw_to_log("---- End of sample ----")


# -------------------------------------------------------------------------
# QUERY 模式：命中 OG 后输出 OG 目录 / counts / summary
# -------------------------------------------------------------------------
def run_query_mode(
    query_arg: str,
    script_dir: Path,
    out_base: Path,
    exact_map: dict[str, set[str]],
    og2sp2prot: dict[str, dict[str, set[str]]],
    species_order: list[str],
    log: Logger,
) -> None:
    query_ids_raw, mode = resolve_query_as_ids(query_arg, script_dir)
    if not query_ids_raw:
        raise RuntimeError("输入的 protein_id 为空（字符串为空或文件无有效行）。")

    query_ids = [normalize_pid_keep_dot(x) for x in query_ids_raw]
    query_ids = [x for x in query_ids if x]

    log.info(f"Mode: {mode}")
    if mode == "file":
        log.info(f"FILE_INPUT_HAS_HEADER={FILE_INPUT_HAS_HEADER}（QUERY 文件第一行将{'被跳过' if FILE_INPUT_HAS_HEADER else '参与匹配'}）")
    log.info(f"Query IDs: total={len(query_ids)}")

    multi_og_counter: dict[str, int] = {}

    hit_ogs: set[str] = set()
    hit_records: set[tuple[str, str, str]] = set()

    for q in query_ids:
        og = match_pid_to_single_og_exact(q, exact_map, multi_og_counter)
        if og is None:
            continue
        hit_ogs.add(og)

        # 记录这个 query 在哪些物种列里出现（避免乱写 Species）
        sp2prot = og2sp2prot.get(og, {})
        for sp, prot_set in sp2prot.items():
            if q in prot_set:
                hit_records.add((q, og, sp))

    if not hit_ogs:
        raise RuntimeError("没有命中任何 OG。")

    out_base.mkdir(parents=True, exist_ok=True)

    hits_summary_path = out_base / HITS_SUMMARY_NAME
    with hits_summary_path.open("w", encoding="utf-8") as w:
        for q, og, sp in sorted(hit_records, key=lambda x: (x[1], x[2], x[0])):
            w.write(f"{q}\t{og}\t{sp}\n")

    # 写 OG/<Species>.txt：写的是该 OG 在该物种下的所有成员（和原脚本一致）
    for og in sorted(hit_ogs):
        og_dir = out_base / og
        og_dir.mkdir(parents=True, exist_ok=True)

        sp2prot = og2sp2prot.get(og, {})
        for sp, prot_set in sp2prot.items():
            sp_fname = sanitize_filename(sp) + ".txt"
            out_path = og_dir / sp_fname
            prot_ids = sorted(list(prot_set))
            with out_path.open("w", encoding="utf-8") as w:
                for pid in prot_ids:
                    w.write(f"{pid}\n")

    counts_matrix_path = out_base / COUNTS_MATRIX_NAME
    with counts_matrix_path.open("w", encoding="utf-8") as w:
        w.write("OG")
        for sp in species_order:
            w.write("\t" + sp)
        w.write("\n")
        for og in sorted(hit_ogs):
            w.write(og)
            for sp in species_order:
                cnt = len(og2sp2prot.get(og, {}).get(sp, set()))
                w.write(f"\t{cnt}")
            w.write("\n")

    log.ok(f"Written: {hits_summary_path}")
    log.ok(f"Written: {counts_matrix_path}")
    log.ok(f"OG folders under: {out_base}")

    if multi_og_counter:
        total = sum(multi_og_counter.values())
        uniq = len(multi_og_counter)
        log.warn(f"QUERY mode: multi-OG exact-id conflicts encountered: total_hits={total}, unique_ids={uniq}. (No per-id spam.)")
        log.raw_to_log("---- QUERY multi-OG conflict sample (pid -> times) ----")
        for pid, times in list(sorted(multi_og_counter.items(), key=lambda x: (-x[1], x[0])))[:50]:
            log.raw_to_log(f"{pid}\t{times}")
        log.raw_to_log("---- End of sample ----")


def main() -> int:
    script_path = Path(__file__).resolve()
    script_dir = script_path.parent

    og_tsv = script_dir / ORTHO_GROUPS_TSV_REL
    out_base = script_dir / OUT_BASE_REL

    log_path = out_base / "logs" / LOG_NAME
    log = Logger(log_path)

    parser = argparse.ArgumentParser(
        prog="gene_hit_og.py",
        add_help=True,
        usage="python scripts/gene_hit_og.py [-p FILE] [QUERY]",
        description="OG hit extractor based on OrthoFinder Orthogroups.tsv (+ optional -p file clustering by OG).",
    )
    parser.add_argument(
        "-p",
        dest="preprocess_file",
        default=None,
        help="输入一个 TSV：新增/覆盖 OG 列并按 OG 聚类排序输出（不删减任何内容）；未命中行 OG=NA 且集中到末尾；可额外输出 XLSX。",
    )
    parser.add_argument(
        "query",
        nargs="?",
        default=None,
        help="protein_id 或包含 protein_id 的文件路径（省略则只执行 -p 模式）。",
    )
    args = parser.parse_args()

    try:
        if not og_tsv.exists():
            raise RuntimeError(f"找不到 Orthogroups.tsv：{ORTHO_GROUPS_TSV_REL}\n脚本目录：{script_dir}")

        log.info(f"Script: {script_path.name}")
        log.info(f"Orthogroups.tsv (rel): {ORTHO_GROUPS_TSV_REL}")
        log.info(f"Output base (rel): {OUT_BASE_REL}")
        log.info(f"FILE_INPUT_HAS_HEADER={FILE_INPUT_HAS_HEADER}")
        log.info(f"WRITE_XLSX_FOR_P={WRITE_XLSX_FOR_P}  FRESHEN_FACTOR={FRESHEN_FACTOR}")
        log.info("ID rule: if contains '|', take substring after last '|'; DO NOT modify anything after '.' (exact match only).")
        log.info(f"Log file: {log_path}")

        exact_map, og2sp2prot, species_order, _ = load_maps_from_orthogroups(og_tsv, log)
        log.info(f"Loaded: species={len(species_order)}  indexed_ids={len(exact_map)}  ogs={len(og2sp2prot)}")

        if args.preprocess_file is not None:
            in_fp = resolve_existing_file(args.preprocess_file, script_dir)
            if in_fp is None:
                raise RuntimeError(f"-p 指定的文件不存在或不可读：{args.preprocess_file}")

            out_dir = out_base / "preprocessed"
            base_name = Path(args.preprocess_file).name
            out_tsv = out_dir / (base_name + ".ogcluster.tsv")
            out_xlsx = out_dir / (base_name + ".ogcluster.xlsx")

            preprocess_file_cluster_by_og(
                in_file=in_fp,
                out_tsv=out_tsv,
                out_xlsx=out_xlsx,
                exact_map=exact_map,
                log=log,
            )
            log.ok(f"Preprocessed TSV:  {out_tsv}")
            if WRITE_XLSX_FOR_P:
                log.ok(f"Preprocessed XLSX: {out_xlsx}")

        if args.query is not None:
            run_query_mode(
                query_arg=args.query,
                script_dir=script_dir,
                out_base=out_base,
                exact_map=exact_map,
                og2sp2prot=og2sp2prot,
                species_order=species_order,
                log=log,
            )

        if args.preprocess_file is None and args.query is None:
            raise RuntimeError("你没有提供任何输入：请给 QUERY 或使用 -p FILE。")

        return 0

    except Exception as e:
        log.error(str(e))
        raise
    finally:
        log.close()


if __name__ == "__main__":
    raise SystemExit(main())


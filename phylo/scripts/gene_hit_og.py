#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gene_hit_og.py

模式A（默认，位置参数 QUERY）：
- QUERY 可以是 protein_id 字符串，也可以是“包含 protein_id 的文件”
- 从 all_pep2cds_resolved.tsv 中找命中的 OG
- 输出：
  1) OUT_BASE/OG_hits_summary.tsv：query_protein_id \t OG \t Species（无表头）
  2) OUT_BASE/OG_species_counts.tsv：行=OG，列=Species，值=该 OG 在该物种下 protein_id 数量
  3) OUT_BASE/<OG>/<Species>.txt：每行一个 protein_id（无表头）

模式B（-p FILE）：
- 接收一个 TSV 文件
- 不删减、不修改任何原始内容（保留所有列与值），仅：
  1) 新增/覆盖第一列 OG
  2) 按 OG 聚类排序（同一 OG 的行相邻）
  3) 无法匹配 OG 的行放到文件末尾，且 OG=NA
- 列规则（按皇上要求）：
  - 如果输入文件第一列是 OG，则 protein_id 在第二列
  - 如果输入文件第一列不是 OG，则 protein_id 在第一列
- 输出：
  - TSV：OUT_BASE/preprocessed/<输入文件名>.ogcluster.tsv
  - XLSX：OUT_BASE/preprocessed/<输入文件名>.ogcluster.xlsx
    * 第一列 OG 的单元格背景色按 OG 区分（NA 使用单独颜色）
    * 颜色来自御用20色，并进行“清新化”处理

运行方式（你在 phylo/ 下运行）：
  python scripts/gene_hit_og.py EVM0019448
  python scripts/gene_hit_og.py input/ids.txt
  python scripts/gene_hit_og.py -p input/table.tsv
"""

from __future__ import annotations

import sys
import argparse
from pathlib import Path
from collections import defaultdict


# =============================================================================
# 顶部参数区（只改这里就行；全部为“相对脚本所在目录”的相对路径）
# =============================================================================

# 文件表头开关：
# True：凡是“输入是文件”（QUERY 文件、-p 文件），默认第一行是表头
# False：文件无表头（第一行当数据）
FILE_INPUT_HAS_HEADER = True

MAP_TSV_REL = "../results/publish/aphylo_ready/all_pep2cds_resolved.tsv"
OUT_BASE_REL = "../results/out_og_hits"

HITS_SUMMARY_NAME = "OG_hits_summary.tsv"
COUNTS_MATRIX_NAME = "OG_species_counts.tsv"

DEDUP_PROTEIN_IDS = True

# -p 模式 XLSX 输出开关
WRITE_XLSX_FOR_P = True

# 皇上御用20色（HEX）
ROYAL_20_HEX = [
    "#F79D93", "#95C8F2", "#F6CD96", "#9FD5CB", "#F6C6E7",
    "#A99BEF", "#F5B07E", "#A7D9F7", "#F3A6C9", "#B6E0B6",
    "#FBE3A8", "#8FE3F2", "#F4B1A8", "#8FB1F2", "#F0D07A",
    "#7FD3C8", "#F7B6D2", "#C3A3F5", "#D6E7B5", "#9E90E6",
]

# “清新化”系数：把颜色与白色混合（0~1）
# 数值越大越浅、越清新；建议 0.35~0.55
FRESHEN_FACTOR = 0.45

# NA 的背景色（清淡灰）
NA_FILL_HEX = "#F2F2F2"

# =============================================================================


def eprint(msg: str) -> None:
    print(msg, file=sys.stderr)


def sanitize_filename(name: str) -> str:
    bad = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    for ch in bad:
        name = name.replace(ch, "_")
    name = name.strip().replace(" ", "_")
    return name


def strip_db_prefix(x: str) -> str:
    if "|" in x:
        return x.split("|", 1)[1]
    return x


def clean_cell(x: str) -> str:
    return x.strip().strip('"').strip("'")


def split_query_tokens(raw: str) -> list[str]:
    raw = raw.strip()
    if not raw:
        return []
    tmp = raw.replace(",", " ")
    toks = [t.strip() for t in tmp.split() if t.strip()]
    return toks


def base_id(pid: str) -> str:
    if "." in pid:
        return pid.split(".", 1)[0]
    return pid


def resolve_existing_file(arg: str, script_dir: Path) -> Path | None:
    p_cwd = Path(arg)
    if p_cwd.exists() and p_cwd.is_file():
        return p_cwd
    p_script = script_dir / arg
    if p_script.exists() and p_script.is_file():
        return p_script
    return None


def read_header_and_col_index(header_line: str) -> dict[str, int]:
    cols = header_line.rstrip("\n").split("\t")
    idx = {c: i for i, c in enumerate(cols)}
    required = ["OG", "Species", "protein_id"]
    missing = [c for c in required if c not in idx]
    if missing:
        raise RuntimeError(f"映射表缺少必要列：{missing}。实际列名：{cols}")
    return idx


def load_og_maps(map_tsv: Path) -> tuple[dict[str, set[str]], dict[str, set[str]]]:
    exact_map: dict[str, set[str]] = defaultdict(set)
    base_map: dict[str, set[str]] = defaultdict(set)

    with map_tsv.open("r", encoding="utf-8", errors="strict") as f:
        header = f.readline()
        if not header:
            raise RuntimeError("映射表是空文件。")
        col = read_header_and_col_index(header)
        og_i = col["OG"]
        pid_i = col["protein_id"]

        for line_no, line in enumerate(f, start=2):
            line = line.rstrip("\n")
            if not line:
                continue
            parts = line.split("\t")
            if len(parts) <= max(og_i, pid_i):
                raise RuntimeError(f"映射表第 {line_no} 行列数不足：{line}")

            og = parts[og_i]
            pid = strip_db_prefix(parts[pid_i])

            exact_map[pid].add(og)
            base_map[base_id(pid)].add(og)

    return exact_map, base_map


def match_pid_to_single_og(
    query_pid: str,
    exact_map: dict[str, set[str]],
    base_map: dict[str, set[str]],
) -> str | None:
    q = strip_db_prefix(clean_cell(query_pid))
    if not q:
        return None

    if "." in q:
        ogs = exact_map.get(q, set())
        if len(ogs) == 1:
            return next(iter(ogs))
        if len(ogs) > 1:
            chosen = sorted(ogs)[0]
            eprint(f"[WARN] protein_id '{q}' 命中多个 OG：{sorted(ogs)}；将使用 {chosen}")
            return chosen

        ogs2 = base_map.get(base_id(q), set())
        if len(ogs2) == 1:
            return next(iter(ogs2))
        if len(ogs2) > 1:
            chosen = sorted(ogs2)[0]
            eprint(f"[WARN] protein_id(base) '{base_id(q)}' 命中多个 OG：{sorted(ogs2)}；将使用 {chosen}")
            return chosen
        return None

    ogs = base_map.get(q, set())
    if len(ogs) == 1:
        return next(iter(ogs))
    if len(ogs) > 1:
        chosen = sorted(ogs)[0]
        eprint(f"[WARN] query '{q}'（无后缀）命中多个 OG：{sorted(ogs)}；将使用 {chosen}")
        return chosen

    ogs2 = exact_map.get(q, set())
    if len(ogs2) == 1:
        return next(iter(ogs2))
    if len(ogs2) > 1:
        chosen = sorted(ogs2)[0]
        eprint(f"[WARN] query '{q}' 命中多个 OG：{sorted(ogs2)}；将使用 {chosen}")
        return chosen

    return None


def resolve_query_as_ids(arg1: str, script_dir: Path) -> tuple[list[str], str]:
    fp = resolve_existing_file(arg1, script_dir)
    if fp is not None:
        ids: list[str] = []
        with fp.open("r", encoding="utf-8", errors="strict") as f:
            first = True
            for line in f:
                if first and FILE_INPUT_HAS_HEADER:
                    first = False
                    continue
                first = False
                line = line.strip()
                if not line:
                    continue
                ids.extend(split_query_tokens(line))
        return ids, "file"
    return split_query_tokens(arg1), "string"


def hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
    h = hex_color.strip().lstrip("#")
    if len(h) != 6:
        raise RuntimeError(f"非法 HEX 颜色：{hex_color}")
    r = int(h[0:2], 16)
    g = int(h[2:4], 16)
    b = int(h[4:6], 16)
    return r, g, b


def rgb_to_hex(r: int, g: int, b: int) -> str:
    r = max(0, min(255, int(r)))
    g = max(0, min(255, int(g)))
    b = max(0, min(255, int(b)))
    return f"#{r:02X}{g:02X}{b:02X}"


def freshen_hex(hex_color: str, factor: float) -> str:
    """
    将颜色与白色混合：new = color*(1-factor) + white*factor
    factor 越大越浅更清新
    """
    r, g, b = hex_to_rgb(hex_color)
    r2 = r * (1 - factor) + 255 * factor
    g2 = g * (1 - factor) + 255 * factor
    b2 = b * (1 - factor) + 255 * factor
    return rgb_to_hex(r2, g2, b2)


def build_og_fill_map(ogs_in_order: list[str]) -> dict[str, str]:
    """
    为 OG 分配“清新化后的”背景色（HEX）。
    同一 OG 同色；超过 20 个则循环使用。
    """
    base = [freshen_hex(c, FRESHEN_FACTOR) for c in ROYAL_20_HEX]
    m: dict[str, str] = {}
    k = 0
    for og in ogs_in_order:
        if og == "NA":
            continue
        if og in m:
            continue
        m[og] = base[k % len(base)]
        k += 1
    return m


def write_xlsx_with_og_colors(xlsx_path: Path, rows: list[list[str]], has_header: bool) -> None:
    """
    rows：完整表格内容（已经包含 OG 列；包含表头行则 rows[0] 是表头）
    仅对第一列 OG 单元格填充背景色（表头行不填充）。
    """
    # 延迟导入，避免非 -p 模式无意义加载
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "OG_cluster"

    # 写入数据
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # 收集 OG 顺序（用于配色一致）
    data_start = 2 if has_header else 1
    ogs_seen: list[str] = []
    for r_idx in range(data_start, len(rows) + 1):
        og = str(ws.cell(row=r_idx, column=1).value or "").strip()
        if not og:
            og = "NA"
        if og not in ogs_seen:
            ogs_seen.append(og)

    og2hex = build_og_fill_map(ogs_seen)

    # 对第一列 OG 着色
    for r_idx in range(data_start, len(rows) + 1):
        og = str(ws.cell(row=r_idx, column=1).value or "").strip()
        if not og:
            og = "NA"
        if og == "NA":
            hex_color = NA_FILL_HEX
        else:
            hex_color = og2hex.get(og, NA_FILL_HEX)

        # openpyxl 的颜色用 ARGB（FF + RRGGBB）
        argb = "FF" + hex_color.lstrip("#").upper()
        fill = PatternFill(patternType="solid", fgColor=argb)
        ws.cell(row=r_idx, column=1).fill = fill

    # 简单调一下列宽（不追求花哨）
    max_cols = max(len(r) for r in rows) if rows else 1
    for c in range(1, max_cols + 1):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 18 if c == 1 else 22

    wb.save(xlsx_path)


def preprocess_file_cluster_by_og(
    in_file: Path,
    out_tsv: Path,
    out_xlsx: Path,
    exact_map: dict[str, set[str]],
    base_map: dict[str, set[str]],
) -> None:
    out_tsv.parent.mkdir(parents=True, exist_ok=True)

    rows_for_xlsx: list[list[str]] = []
    has_header = FILE_INPUT_HAS_HEADER

    with in_file.open("r", encoding="utf-8", errors="strict") as f:
        first_line = f.readline()
        if not first_line:
            raise RuntimeError(f"-p 输入文件为空：{in_file}")

        header_fields: list[str] | None = None
        og_already_first = False

        if FILE_INPUT_HAS_HEADER:
            header_fields = first_line.rstrip("\n").split("\t")
            if len(header_fields) == 0:
                raise RuntimeError(f"-p 输入文件表头为空：{in_file}")
            og_already_first = (header_fields[0].strip().lower().lstrip("\ufeff") == "og")
        else:
            # 无表头：不做猜测，直接按列规则：若第一列是 OG（极少见），就用第二列，否则第一列
            data0 = first_line.rstrip("\n").split("\t")
            if len(data0) == 0:
                raise RuntimeError(f"-p 输入文件首行为空：{in_file}")
            og_already_first = (clean_cell(data0[0]).strip().lower() == "og")

        # 关键列选择（按皇上规则）
        id_col = 1 if og_already_first else 0

        groups: dict[str, list[list[str]]] = {}
        og_order: list[str] = []
        unmatched_rows: list[list[str]] = []

        def add_to_group(og: str, row_fields: list[str]) -> None:
            if og not in groups:
                groups[og] = []
                og_order.append(og)
            groups[og].append(row_fields)

        def process_row(fields: list[str], line_no: int) -> None:
            if id_col >= len(fields):
                raise RuntimeError(f"-p 输入文件第 {line_no} 行列数不足：{fields}")

            pid = fields[id_col]
            og = match_pid_to_single_og(pid, exact_map, base_map)
            og_val = og if og is not None else "NA"

            if og_already_first:
                new_fields = fields[:]
                new_fields[0] = og_val
            else:
                new_fields = [og_val] + fields

            if og is None:
                unmatched_rows.append(new_fields)
            else:
                add_to_group(og, new_fields)

        # 读取并处理数据
        if FILE_INPUT_HAS_HEADER:
            header_out = header_fields if og_already_first else (["OG"] + header_fields)
            rows_for_xlsx.append(header_out)

            for line_no, line in enumerate(f, start=2):
                line = line.rstrip("\n")
                if line == "":
                    continue
                process_row(line.split("\t"), line_no)
        else:
            process_row(first_line.rstrip("\n").split("\t"), 1)
            for line_no, line in enumerate(f, start=2):
                line = line.rstrip("\n")
                if line == "":
                    continue
                process_row(line.split("\t"), line_no)

        # 组织输出顺序：按首次出现的 OG 依次输出，最后输出 unmatched
        for og in og_order:
            for row in groups[og]:
                rows_for_xlsx.append(row)
        for row in unmatched_rows:
            rows_for_xlsx.append(row)

    # 写 TSV
    with out_tsv.open("w", encoding="utf-8") as w:
        for row in rows_for_xlsx:
            w.write("\t".join(row) + "\n")

    # 写 XLSX
    if WRITE_XLSX_FOR_P:
        write_xlsx_with_og_colors(out_xlsx, rows_for_xlsx, has_header=has_header)


def run_query_mode(query_arg: str, script_dir: Path, map_tsv: Path, out_base: Path) -> None:
    query_ids_raw, mode = resolve_query_as_ids(query_arg, script_dir)
    if not query_ids_raw:
        raise RuntimeError("输入的 protein_id 为空（字符串为空或文件无有效行）。")

    query_ids = [strip_db_prefix(x) for x in query_ids_raw]

    exact_ids = set()
    prefix_ids = set()
    for q in query_ids:
        q = q.strip()
        if not q:
            continue
        if "." in q:
            exact_ids.add(q)
        else:
            prefix_ids.add(q)

    eprint(f"[INFO] Mode: {mode}")
    if mode == "file":
        eprint(f"[INFO] FILE_INPUT_HAS_HEADER={FILE_INPUT_HAS_HEADER}（QUERY 文件第一行将{'被跳过' if FILE_INPUT_HAS_HEADER else '参与匹配'}）")
    eprint(f"[INFO] Query IDs: total={len(query_ids)} exact={len(exact_ids)} prefix={len(prefix_ids)}")

    hit_ogs: set[str] = set()
    hit_records: set[tuple[str, str, str]] = set()

    with map_tsv.open("r", encoding="utf-8", errors="strict") as f:
        header = f.readline()
        if not header:
            raise RuntimeError("映射表是空文件。")
        col = read_header_and_col_index(header)

        og_i = col["OG"]
        sp_i = col["Species"]
        pid_i = col["protein_id"]

        for line_no, line in enumerate(f, start=2):
            line = line.rstrip("\n")
            if not line:
                continue
            parts = line.split("\t")
            if len(parts) <= max(og_i, sp_i, pid_i):
                raise RuntimeError(f"映射表第 {line_no} 行列数不足：{line}")

            og = parts[og_i]
            sp = parts[sp_i]
            pid = strip_db_prefix(parts[pid_i])
            pid_base = base_id(pid)

            if pid in exact_ids:
                hit_ogs.add(og)
                hit_records.add((pid, og, sp))
            if pid_base in prefix_ids:
                hit_ogs.add(og)
                hit_records.add((pid_base, og, sp))

    if not hit_ogs:
        raise RuntimeError("没有命中任何 OG。")

    og2sp2prot: dict[str, dict[str, set[str] | list[str]]] = defaultdict(
        lambda: defaultdict(set if DEDUP_PROTEIN_IDS else list)
    )
    species_order: list[str] = []

    def add_species_once(s: str) -> None:
        if s not in species_order:
            species_order.append(s)

    with map_tsv.open("r", encoding="utf-8", errors="strict") as f:
        header = f.readline()
        col = read_header_and_col_index(header)
        og_i = col["OG"]
        sp_i = col["Species"]
        pid_i = col["protein_id"]

        for line_no, line in enumerate(f, start=2):
            line = line.rstrip("\n")
            if not line:
                continue
            parts = line.split("\t")
            if len(parts) <= max(og_i, sp_i, pid_i):
                raise RuntimeError(f"映射表第 {line_no} 行列数不足：{line}")

            og = parts[og_i]
            if og not in hit_ogs:
                continue

            sp = parts[sp_i]
            pid = strip_db_prefix(parts[pid_i])

            add_species_once(sp)

            if DEDUP_PROTEIN_IDS:
                og2sp2prot[og][sp].add(pid)  # type: ignore[union-attr]
            else:
                og2sp2prot[og][sp].append(pid)  # type: ignore[union-attr]

    out_base.mkdir(parents=True, exist_ok=True)

    hits_summary_path = out_base / HITS_SUMMARY_NAME
    with hits_summary_path.open("w", encoding="utf-8") as w:
        for q, og, sp in sorted(hit_records, key=lambda x: (x[1], x[2], x[0])):
            w.write(f"{q}\t{og}\t{sp}\n")

    for og in sorted(hit_ogs):
        og_dir = out_base / og
        og_dir.mkdir(parents=True, exist_ok=True)

        sp2prot = og2sp2prot.get(og, {})
        for sp, prot_container in sp2prot.items():
            sp_fname = sanitize_filename(sp) + ".txt"
            out_path = og_dir / sp_fname

            if DEDUP_PROTEIN_IDS:
                prot_ids = sorted(list(prot_container))  # type: ignore[arg-type]
            else:
                prot_ids = list(prot_container)  # type: ignore[arg-type]

            with out_path.open("w", encoding="utf-8") as w:
                for pid in prot_ids:
                    w.write(f"{pid}\n")

    counts_matrix_path = out_base / COUNTS_MATRIX_NAME
    with counts_matrix_path.open("w", encoding="utf-8") as w:
        w.write("OG")
        for sp in species_order:
            w.write("\t" + sp)
        w.write("\n")

        for og in sorted(hit_ogs):
            w.write(og)
            for sp in species_order:
                prot_container = og2sp2prot.get(og, {}).get(sp, set() if DEDUP_PROTEIN_IDS else [])
                cnt = len(prot_container)  # type: ignore[arg-type]
                w.write(f"\t{cnt}")
            w.write("\n")

    eprint(f"[OK] Written: {hits_summary_path}")
    eprint(f"[OK] Written: {counts_matrix_path}")
    eprint(f"[OK] OG folders under: {out_base}")


def main() -> int:
    script_path = Path(__file__).resolve()
    script_dir = script_path.parent

    map_tsv = script_dir / MAP_TSV_REL
    out_base = script_dir / OUT_BASE_REL

    parser = argparse.ArgumentParser(
        prog="gene_hit_og.py",
        add_help=True,
        usage="python scripts/gene_hit_og.py [-p FILE] [QUERY]",
        description="OG hit extractor (+ optional -p file clustering by OG).",
    )
    parser.add_argument(
        "-p",
        dest="preprocess_file",
        default=None,
        help="输入一个文件：新增/覆盖 OG 列并按 OG 聚类排序输出（不删减任何内容），并额外输出 XLSX（若开关开启）。",
    )
    parser.add_argument(
        "query",
        nargs="?",
        default=None,
        help="protein_id 或包含 protein_id 的文件路径（省略则只执行 -p 模式）。",
    )

    args = parser.parse_args()

    if not map_tsv.exists():
        raise RuntimeError(f"找不到映射表：{MAP_TSV_REL}\n脚本目录：{script_dir}")

    eprint(f"[INFO] Script: {script_path.name}")
    eprint(f"[INFO] Map TSV (rel): {MAP_TSV_REL}")
    eprint(f"[INFO] Output base (rel): {OUT_BASE_REL}")
    eprint(f"[INFO] FILE_INPUT_HAS_HEADER={FILE_INPUT_HAS_HEADER}")
    eprint(f"[INFO] WRITE_XLSX_FOR_P={WRITE_XLSX_FOR_P}  FRESHEN_FACTOR={FRESHEN_FACTOR}")

    exact_map, base_map = load_og_maps(map_tsv)

    if args.preprocess_file is not None:
        in_fp = resolve_existing_file(args.preprocess_file, script_dir)
        if in_fp is None:
            raise RuntimeError(f"-p 指定的文件不存在或不可读：{args.preprocess_file}")

        out_dir = out_base / "preprocessed"
        base_name = Path(args.preprocess_file).name

        out_tsv = out_dir / (base_name + ".ogcluster.tsv")
        out_xlsx = out_dir / (base_name + ".ogcluster.xlsx")

        preprocess_file_cluster_by_og(
            in_file=in_fp,
            out_tsv=out_tsv,
            out_xlsx=out_xlsx,
            exact_map=exact_map,
            base_map=base_map,
        )
        eprint(f"[OK] Preprocessed TSV:  {out_tsv}")
        if WRITE_XLSX_FOR_P:
            eprint(f"[OK] Preprocessed XLSX: {out_xlsx}")

    if args.query is not None:
        run_query_mode(
            query_arg=args.query,
            script_dir=script_dir,
            map_tsv=map_tsv,
            out_base=out_base,
        )

    if args.preprocess_file is None and args.query is None:
        raise RuntimeError("你没有提供任何输入：请给 QUERY 或使用 -p FILE。")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as e:
        eprint(f"[ERROR] {e}")
        raise

#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
09_publish_results.py —— RNA-seq vNext 发布与装订脚本（含可选 xlsx + GO/KEGG 宽表）

功能定位（不做新统计，只做“装订 + 验收 + 说明 + 可选 xlsx”）：
  1. 从 05 / 06 / 08 模块收集顶刊需要的表格，复制到：
       results/09_publish/source_data/
     目录结构：
       - matrix/
       - deg/<contrast>/
       - background/
       - enrich/<label>/
  2. 生成两个“验收单”：
       - manifest.tsv      —— 所有已打包文件的清单
       - publish_check.tsv —— 各模块缺失文件情况
  3. 生成 METHODS_README.txt 骨架：
       - 填入 config.yaml 中的关键统计口径与阈值
       - 提供样本数量 / 分组情况等信息
  4. 若 config.publish.xlsx = true，则为每个 contrast 生成一个 xlsx：
       results/09_publish/<contrast>.xlsx
     xlsx 仅从 source_data 读现有 TSV：
       - sheet：DEG_all / GO_sig_up / GO_sig_down / KEGG_all / background_meta
       - 宽表：GO_summary（GO all/up/down 汇总），KEGG_summary（KEGG all/up/down 汇总）
     不改变任何统计值与 TSV 表头，严格遵守 vNext 约定。

使用方式：
  在项目根目录运行（不接受命令行参数）：
    python scripts/09_publish_results.py

依赖：
  - PyYAML（yaml）
  - Python 3.8+
  - openpyxl（可选；仅在 publish.xlsx=true 且成功导入 openpyxl 时使用）
"""

from __future__ import annotations

import sys
import shutil
import logging
from pathlib import Path
from typing import Dict, List, Any

import yaml
import csv

# ======================= 可在此处手动调整的参数 =======================

# 配置文件路径
CONFIG_FILE: str = "config.yaml"

# manifest.tsv 列名
MANIFEST_COLUMNS = ["path", "module", "label", "description"]

# publish_check.tsv 列名
CHECK_COLUMNS = ["module", "label", "status", "detail"]

# ====================================================================


def setup_logging(level: str = "INFO") -> None:
    """初始化日志系统，打印到标准输出，带时间戳。"""
    numeric_level = getattr(logging, level.upper(), logging.INFO)
    logging.basicConfig(
        level=numeric_level,
        format="%(asctime)s [09_publish] [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def load_config(path: Path) -> Dict[str, Any]:
    """
    读取 YAML 配置文件。
    """
    if not path.is_file():
        logging.error("找不到配置文件：%s", path)
        sys.exit(1)
    with path.open("r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    logging.info("已加载配置文件：%s", path)
    return cfg


def ensure_dir(p: Path) -> None:
    """若目录不存在则创建目录（包含父目录）。"""
    p.mkdir(parents=True, exist_ok=True)


def copy_file(
    src: Path,
    dst: Path,
    module: str,
    label: str,
    description: str,
    manifest_rows: List[Dict[str, str]],
    source_data_root: Path,
) -> None:
    """
    复制单个文件，并将其记录到 manifest_rows。
    """
    ensure_dir(dst.parent)
    shutil.copy2(src, dst)
    rel_path = dst.relative_to(source_data_root)
    manifest_rows.append(
        {
            "path": str(rel_path),
            "module": module,
            "label": label if label else "NA",
            "description": description,
        }
    )
    logging.info("已复制文件：%s -> %s", src, dst)


def write_tsv(path: Path, rows: List[Dict[str, Any]], fieldnames: List[str]) -> None:
    """
    写出 TSV 文件。
    """
    ensure_dir(path.parent)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=fieldnames,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()
        for row in rows:
            safe_row = {k: row.get(k, "") for k in fieldnames}
            writer.writerow(safe_row)
    logging.info("已写出 TSV：%s（%d 行+表头）", path, len(rows))


def read_tsv_header_and_rows(path: Path) -> List[Dict[str, str]]:
    """
    简单读取 TSV（带表头），全部转为字符串。
    """
    rows: List[Dict[str, str]] = []
    with path.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            rows.append({k: (v if v is not None else "") for k, v in row.items()})
    return rows


def get_contrasts_list(cfg: Dict[str, Any]) -> List[str]:
    """
    从 data/contrasts.tsv 读取 contrast 列，得到对比列表。
    """
    data_cfg = cfg.get("data", {})
    contrasts_path = Path(data_cfg.get("contrasts_tsv", "data/contrasts.tsv"))
    if not contrasts_path.is_file():
        logging.warning("未找到 contrasts.tsv（%s），将不基于 contrasts 表生成对比列表。", contrasts_path)
        return []

    rows = read_tsv_header_and_rows(contrasts_path)
    contrasts: List[str] = []
    for row in rows:
        c = row.get("contrast", "").strip()
        if c:
            contrasts.append(c)
    uniq_contrasts = sorted(set(contrasts))
    logging.info("从 contrasts.tsv 读取到对比数量：%d", len(uniq_contrasts))
    return uniq_contrasts


def summarize_samples(cfg: Dict[str, Any]) -> Dict[str, Any]:
    """
    从 data/samples.tsv 中统计样本数、分组情况，用于 Methods 文本。
    """
    data_cfg = cfg.get("data", {})
    samples_path = Path(data_cfg.get("samples_tsv", "data/samples.tsv"))
    summary = {
        "n_samples": 0,
        "group_counts": {},
        "has_batch": False,
    }
    if not samples_path.is_file():
        logging.warning("未找到 samples.tsv（%s），METHODS 中样本统计将简化。", samples_path)
        return summary

    rows = read_tsv_header_and_rows(samples_path)
    group_counts: Dict[str, int] = {}
    has_batch = False

    for row in rows:
        group = row.get("group", "").strip()
        if group:
            group_counts[group] = group_counts.get(group, 0) + 1
        if "batch" in row and row.get("batch", "").strip():
            has_batch = True

    summary["n_samples"] = len(rows)
    summary["group_counts"] = group_counts
    summary["has_batch"] = has_batch
    logging.info(
        "samples.tsv 统计：总样本数=%d，分组数=%d（%s），batch 列=%s",
        summary["n_samples"],
        len(group_counts),
        ", ".join(f"{k}:{v}" for k, v in group_counts.items()) if group_counts else "无",
        "存在" if has_batch else "不存在或全空",
    )
    return summary


def build_methods_readme_text(
    cfg: Dict[str, Any],
    samples_summary: Dict[str, Any],
    source_data_root: Path,
) -> str:
    """
    构造 METHODS_README.txt 的文本内容。
    """
    project_cfg = cfg.get("project", {})
    reference_cfg = cfg.get("reference", {})
    deg_cfg = cfg.get("deg", {})
    background_cfg = cfg.get("background", {})
    go_cfg = cfg.get("go", {})
    kegg_cfg = cfg.get("kegg", {})
    enrich_cfg = cfg.get("enrich", {})
    gsea_cfg = cfg.get("gsea", {})

    name = project_cfg.get("name", "rnaseq_project")
    species = project_cfg.get("species", "NA")

    n_samples = samples_summary.get("n_samples", 0)
    group_counts = samples_summary.get("group_counts", {})
    has_batch = samples_summary.get("has_batch", False)
    group_desc = ", ".join(f"{g} (n={c})" for g, c in group_counts.items()) if group_counts else "NA"

    deg_lfc = deg_cfg.get("lfc", 1.0)
    deg_fdr = deg_cfg.get("fdr", 0.05)
    use_apeglm = bool(deg_cfg.get("use_apeglm", True))
    independent_filter = bool(deg_cfg.get("independent_filter", True))
    min_samples_per_group = deg_cfg.get("min_samples_per_group", 3)

    background_rule = background_cfg.get("rule", "per_contrast_detectable_AND_annotated")
    background_detectable = background_cfg.get("detectable", "TPM>0_or_count>0_in>=1_sample")

    go_minGS = go_cfg.get("minGS", 10)
    go_maxGS = go_cfg.get("maxGS", 500)
    go_padj_method = go_cfg.get("p_adjust_method", "BH")

    kegg_count_mode = kegg_cfg.get("count_mode", "by_gene")
    kegg_padj_method = kegg_cfg.get("p_adjust_method", "BH")

    enrich_fdr = enrich_cfg.get("fdr", 0.05)

    gsea_enable = bool(gsea_cfg.get("enable", False))
    gsea_score_from = gsea_cfg.get("score_from", "stat")
    gsea_minGS = gsea_cfg.get("minGS", 10)
    gsea_maxGS = gsea_cfg.get("maxGS", 500)
    gsea_padj_method = gsea_cfg.get("p_adjust_method", "BH")
    gsea_fdr = gsea_cfg.get("fdr", 0.25)
    gsea_ontologies = gsea_cfg.get("ontologies", ["BP"])

    matrix_dir_rel = Path("matrix")
    deg_dir_rel = Path("deg")
    enrich_dir_rel = Path("enrich")
    background_dir_rel = Path("background")

    lines: List[str] = []

    lines.append(f"Project: {name}")
    lines.append(f"Species: {species}")
    lines.append("")
    lines.append("1. Samples & Library")
    lines.append("---------------------")
    lines.append(f"- Total number of samples: {n_samples}")
    lines.append(f"- Experimental groups (sample counts): {group_desc}")
    lines.append(f"- Batch column in samples.tsv: {'present' if has_batch else 'absent or not used'}")
    lines.append("- Raw and cleaned FASTQ files were processed by the in-house RNA-seq vNext pipeline (fastp-based QC; see QC tables outside this folder if needed).")
    lines.append("")
    lines.append("2. Quantification")
    lines.append("------------------")
    lines.append("- Quantification was performed using Salmon in quasi-mapping mode with a decoy-aware index.")
    lines.append(f"- Reference genome: {reference_cfg.get('ref_fasta', 'NA')}")
    lines.append(f"- Reference annotation (GFF3/GTF): {reference_cfg.get('ref_gtf', 'NA')}")
    salmon_cfg = reference_cfg.get("salmon", {})
    lines.append(f"- Salmon index directory: {salmon_cfg.get('index_dir', 'NA')}")
    lines.append(f"- Salmon k-mer length: {salmon_cfg.get('kmer_len', 'NA')}")
    lines.append(f"- Salmon libtype: {salmon_cfg.get('libtype', 'A')} (automatic library type detection)")
    lines.append("")
    lines.append("3. Aggregation (tximport and gene-level matrices)")
    lines.append("-----------------------------------------------")
    lines.append("- Transcript-level estimates from Salmon were imported and summarized to the gene level using tximport.")
    lines.append("- Gene-level counts and TPMs are provided in:")
    lines.append(f"  * {matrix_dir_rel/'gene_counts.tsv'}")
    lines.append(f"  * {matrix_dir_rel/'gene_tpm.tsv'}")
    lines.append(f"  * {matrix_dir_rel/'matrix_stats.tsv'} (basic matrix diagnostics, e.g. number of genes, library sizes).")
    tx_cfg = cfg.get("tximport", {})
    lines.append(f"- tximport counts_from_abundance parameter: {tx_cfg.get('counts_from_abundance', 'no')}")
    lines.append("- Gene identifiers (gene_id) were used as the primary keys across all modules; transcript-level IDs were only used internally during aggregation.")
    lines.append("")
    lines.append("4. Differential expression analysis")
    lines.append("------------------------------------")
    lines.append("- Differential expression was performed using DESeq2.")
    lines.append(f"- Design formula included the experimental group, and batch was included when the sample size allowed (min_samples_per_group >= {min_samples_per_group}).")
    lines.append(f"- Log2 fold-change threshold: |log2FC| >= {deg_lfc}")
    lines.append(f"- FDR threshold (Benjamini–Hochberg): {deg_fdr}")
    lines.append(f"- LFC shrinkage: {'apeglm' if use_apeglm else 'not used'}")
    lines.append(f"- Independent filtering: {'enabled' if independent_filter else 'disabled'}")
    lines.append(f"- Per-contrast DEG results and diagnostics are provided under: {deg_dir_rel}/<contrast>/")
    lines.append("  * DEG_all.tsv, DEG_up.tsv, DEG_down.tsv")
    lines.append("  * design.txt (design formula and covariates)")
    lines.append("  * varTop100.list, rle_range.tsv (diagnostics)")
    lines.append("")
    lines.append("5. Enrichment analysis (GO / KEGG ORA and optional GSEA)")
    lines.append("--------------------------------------------------------")
    lines.append("5.1 Background definition")
    lines.append(f"- For each RNA-seq contrast, the enrichment background was defined as:")
    lines.append(f"  * Rule: {background_rule}")
    lines.append(f"  * Detectable criteria: {background_detectable}")
    lines.append(f"- Per-contrast background lists and metadata are provided under: {background_dir_rel}/")
    lines.append("  * <contrast>.list (gene_id of the background universe)")
    lines.append("  * <contrast>.meta.tsv (universe_size, coverage, and sample subset used).")
    lines.append("")
    lines.append("5.2 GO over-representation analysis (ORA)")
    lines.append(f"- GO ORA was performed using a hypergeometric test with {go_padj_method} multiple-testing correction.")
    lines.append(f"- Minimum and maximum gene set size: minGS = {go_minGS}, maxGS = {go_maxGS}.")
    lines.append("- GO terms were tested separately for the three ontologies (BP, CC, MF).")
    lines.append("- For each label (RNA contrast or external tag), by-term tables are provided in:")
    lines.append(f"  * {enrich_dir_rel}/<label>/GO_BP_by_term_(all|up|down).tsv")
    lines.append(f"  * {enrich_dir_rel}/<label>/GO_CC_by_term_(all|up|down).tsv")
    lines.append(f"  * {enrich_dir_rel}/<label>/GO_MF_by_term_(all|up|down).tsv")
    lines.append("- For convenience, GO_sig_all.tsv, GO_sig_up.tsv, and GO_sig_down.tsv provide FDR-filtered summaries across the three ontologies for each label.")
    lines.append(f"- Enrichment FDR cutoff for GO by-term and GO_sig_* tables: {enrich_fdr}.")
    lines.append("")
    lines.append("5.3 KEGG over-representation analysis (ORA)")
    lines.append(f"- KEGG ORA was performed using a hypergeometric test with {kegg_padj_method} multiple-testing correction.")
    lines.append(f"- Gene counts in KEGG terms were computed with count_mode = {kegg_count_mode}.")
    lines.append("- For each label (RNA contrast or external tag), KEGG by-term tables are provided in:")
    lines.append(f"  * {enrich_dir_rel}/<label>/KEGG_by_term_(all|up|down).tsv")
    lines.append(f"- The same enrichment FDR cutoff ({enrich_fdr}) was applied when interpreting KEGG results.")
    lines.append("")
    lines.append("5.4 Gene set enrichment analysis (GSEA) (optional)")
    if gsea_enable:
        lines.append("- GSEA was enabled for RNA-seq contrasts.")
        lines.append(f"- Ranking metric was derived from DEG statistics: score_from = {gsea_score_from}.")
        lines.append(f"- GSEA gene set size range: minGS = {gsea_minGS}, maxGS = {gsea_maxGS}.")
        lines.append(f"- Multiple-testing correction: {gsea_padj_method}, with FDR threshold = {gsea_fdr}.")
        lines.append(f"- GO GSEA ontologies included: {', '.join(gsea_ontologies)}.")
        lines.append(f"- Results are provided (when available) under: {enrich_dir_rel}/<label>/gsea/")
        lines.append("  * GO_gsea.tsv")
        lines.append("  * KEGG_gsea.tsv")
    else:
        lines.append("- GSEA was not enabled in this analysis run.")
    lines.append("")
    lines.append("6. Key tables and source data organisation")
    lines.append("------------------------------------------")
    lines.append(f"- All source data are organised under: {source_data_root}")
    lines.append("- Key subdirectories and their contents:")
    lines.append(f"  * matrix/       : gene-level counts, TPMs, and basic matrix statistics.")
    lines.append(f"  * deg/          : per-contrast DESeq2 outputs and diagnostics.")
    lines.append(f"  * background/   : per-contrast enrichment background lists and metadata.")
    lines.append(f"  * enrich/       : GO/KEGG ORA outputs (and optional GSEA results) for RNA contrasts and external tags.")
    lines.append("- A machine-readable manifest of all files is provided as manifest.tsv.")
    lines.append("- A QC-style summary of missing or incomplete outputs is provided as publish_check.tsv.")
    lines.append("")
    lines.append("7. Notes for reuse")
    lines.append("-------------------")
    lines.append("- All identifiers are reported at the gene_id level, which serves as the primary key across all tables.")
    lines.append("- Any external gene sets (e.g. PSG, CAFE-expanded or contracted families, ATAC-seq peaks mapped to genes) should be mapped to the same gene_id namespace before being passed to the enrichment module.")
    lines.append("- The present pipeline does not generate figures; all results are provided as plain-text tables intended for downstream visualisation and for submission as source data.")

    return "\n".join(lines) + "\n"

# ========================= xlsx 相关工具函数 =========================

def tsv_to_worksheet(ws, tsv_path: Path) -> None:
    """
    将一个 TSV 文件写入到 openpyxl 的 worksheet 中（原样长表写入）。
    """
    with tsv_path.open("r", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter="\t")
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)


def build_go_summary_sheet(
    wb,
    contrast: str,
    source_data_root: Path,
) -> None:
    """
    构建 GO_summary 宽表 sheet：

    数据来源（全部从 source_data/enrich/<contrast>/ 中读取）：
      - GO_BP_by_term_all.tsv / up.tsv / down.tsv
      - GO_CC_by_term_all.tsv / up.tsv / down.tsv
      - GO_MF_by_term_all.tsv / up.tsv / down.tsv

    聚合规则：
      - 以 (term_id, ontology) 为 key，一行一个 term；
      - 为 all/up/down 分别存 p_adjust / gene_count / gene_ids；
      - bg_size/universe_size/min_gs/max_gs 使用任一来源行（通常相同）。

    不做新的统计，仅为 Excel 浏览重组宽表。
    """
    enrich_dir = source_data_root / "enrich" / contrast
    onts = ["BP", "CC", "MF"]
    sets = ["all", "up", "down"]

    # term_key -> 聚合信息
    summary: Dict[tuple, Dict[str, Any]] = {}

    for ont in onts:
        for test_set in sets:
            fname = f"GO_{ont}_by_term_{test_set}.tsv"
            path = enrich_dir / fname
            if not path.is_file():
                continue
            rows = read_tsv_header_and_rows(path)
            for row in rows:
                term_id = row.get("term_id", "").strip()
                term_name = row.get("term_name", "")
                ontology = row.get("ontology", ont)
                if not term_id:
                    continue
                key = (term_id, ontology)
                if key not in summary:
                    summary[key] = {
                        "term_id": term_id,
                        "term_name": term_name,
                        "ontology": ontology,
                        # all/up/down：p_adjust / gene_count / gene_ids
                        "p_adjust_all": "",
                        "p_adjust_up": "",
                        "p_adjust_down": "",
                        "gene_count_all": "",
                        "gene_count_up": "",
                        "gene_count_down": "",
                        "gene_ids_all": "",
                        "gene_ids_up": "",
                        "gene_ids_down": "",
                        # 共享字段
                        "bg_size": row.get("bg_size", ""),
                        "universe_size": row.get("universe_size", ""),
                        "min_gs": row.get("min_gs", ""),
                        "max_gs": row.get("max_gs", ""),
                    }
                dest = summary[key]
                # 根据 test_set 写入对应字段
                if test_set == "all":
                    dest["p_adjust_all"] = row.get("p_adjust", "")
                    dest["gene_count_all"] = row.get("gene_count", "")
                    dest["gene_ids_all"] = row.get("gene_ids", "")
                elif test_set == "up":
                    dest["p_adjust_up"] = row.get("p_adjust", "")
                    dest["gene_count_up"] = row.get("gene_count", "")
                    dest["gene_ids_up"] = row.get("gene_ids", "")
                elif test_set == "down":
                    dest["p_adjust_down"] = row.get("p_adjust", "")
                    dest["gene_count_down"] = row.get("gene_count", "")
                    dest["gene_ids_down"] = row.get("gene_ids", "")
                # 共享字段以最后一次为准（通常一致）
                dest["bg_size"] = row.get("bg_size", dest["bg_size"])
                dest["universe_size"] = row.get("universe_size", dest["universe_size"])
                dest["min_gs"] = row.get("min_gs", dest["min_gs"])
                dest["max_gs"] = row.get("max_gs", dest["max_gs"])

    if not summary:
        logging.info("contrast=%s：未找到 GO by-term 表，跳过 GO_summary 宽表。", contrast)
        return

    ws = wb.create_sheet(title="GO_summary")
    header = [
        "term_id",
        "term_name",
        "ontology",
        "p_adjust_all",
        "p_adjust_up",
        "p_adjust_down",
        "gene_count_all",
        "gene_count_up",
        "gene_count_down",
        "gene_ids_all",
        "gene_ids_up",
        "gene_ids_down",
        "bg_size",
        "universe_size",
        "min_gs",
        "max_gs",
    ]
    ws.append(header)

    # 按 ontology+term_id 排序，方便浏览
    for key in sorted(summary.keys(), key=lambda x: (x[1], x[0])):
        v = summary[key]
        row = [
            v.get("term_id", ""),
            v.get("term_name", ""),
            v.get("ontology", ""),
            v.get("p_adjust_all", ""),
            v.get("p_adjust_up", ""),
            v.get("p_adjust_down", ""),
            v.get("gene_count_all", ""),
            v.get("gene_count_up", ""),
            v.get("gene_count_down", ""),
            v.get("gene_ids_all", ""),
            v.get("gene_ids_up", ""),
            v.get("gene_ids_down", ""),
            v.get("bg_size", ""),
            v.get("universe_size", ""),
            v.get("min_gs", ""),
            v.get("max_gs", ""),
        ]
        ws.append(row)

    logging.info("contrast=%s：已在 xlsx 中添加 GO_summary 宽表（%d 个 term）。", contrast, len(summary))


def build_kegg_summary_sheet(
    wb,
    contrast: str,
    source_data_root: Path,
) -> None:
    """
    构建 KEGG_summary 宽表 sheet：

    数据来源（全部从 source_data/enrich/<contrast>/ 中读取）：
      - KEGG_by_term_all.tsv / up.tsv / down.tsv

    聚合规则：
      - 以 pathway_id 为 key，一行一个 pathway；
      - 为 all/up/down 分别存 p_adjust / gene_count / gene_ids；
      - bg_size/universe_size/min_gs/max_gs/count_mode 使用任一来源行。

    不做新的统计，仅为 Excel 浏览重组宽表。
    """
    enrich_dir = source_data_root / "enrich" / contrast
    sets = ["all", "up", "down"]

    summary: Dict[str, Dict[str, Any]] = {}

    for test_set in sets:
        fname = f"KEGG_by_term_{test_set}.tsv"
        path = enrich_dir / fname
        if not path.is_file():
            continue
        rows = read_tsv_header_and_rows(path)
        for row in rows:
            pathway_id = row.get("pathway_id", "").strip()
            if not pathway_id:
                continue
            term_name = row.get("term_name", "")
            if pathway_id not in summary:
                summary[pathway_id] = {
                    "pathway_id": pathway_id,
                    "term_name": term_name,
                    "p_adjust_all": "",
                    "p_adjust_up": "",
                    "p_adjust_down": "",
                    "gene_count_all": "",
                    "gene_count_up": "",
                    "gene_count_down": "",
                    "gene_ids_all": "",
                    "gene_ids_up": "",
                    "gene_ids_down": "",
                    "bg_size": row.get("bg_size", ""),
                    "universe_size": row.get("universe_size", ""),
                    "min_gs": row.get("min_gs", ""),
                    "max_gs": row.get("max_gs", ""),
                    "count_mode": row.get("count_mode", ""),
                }
            dest = summary[pathway_id]
            if test_set == "all":
                dest["p_adjust_all"] = row.get("p_adjust", "")
                dest["gene_count_all"] = row.get("gene_count", "")
                dest["gene_ids_all"] = row.get("gene_ids", "")
            elif test_set == "up":
                dest["p_adjust_up"] = row.get("p_adjust", "")
                dest["gene_count_up"] = row.get("gene_count", "")
                dest["gene_ids_up"] = row.get("gene_ids", "")
            elif test_set == "down":
                dest["p_adjust_down"] = row.get("p_adjust", "")
                dest["gene_count_down"] = row.get("gene_count", "")
                dest["gene_ids_down"] = row.get("gene_ids", "")

            dest["bg_size"] = row.get("bg_size", dest["bg_size"])
            dest["universe_size"] = row.get("universe_size", dest["universe_size"])
            dest["min_gs"] = row.get("min_gs", dest["min_gs"])
            dest["max_gs"] = row.get("max_gs", dest["max_gs"])
            dest["count_mode"] = row.get("count_mode", dest["count_mode"])

    if not summary:
        logging.info("contrast=%s：未找到 KEGG by-term 表，跳过 KEGG_summary 宽表。", contrast)
        return

    ws = wb.create_sheet(title="KEGG_summary")
    header = [
        "pathway_id",
        "term_name",
        "p_adjust_all",
        "p_adjust_up",
        "p_adjust_down",
        "gene_count_all",
        "gene_count_up",
        "gene_count_down",
        "gene_ids_all",
        "gene_ids_up",
        "gene_ids_down",
        "bg_size",
        "universe_size",
        "min_gs",
        "max_gs",
        "count_mode",
    ]
    ws.append(header)

    for pid in sorted(summary.keys()):
        v = summary[pid]
        row = [
            v.get("pathway_id", ""),
            v.get("term_name", ""),
            v.get("p_adjust_all", ""),
            v.get("p_adjust_up", ""),
            v.get("p_adjust_down", ""),
            v.get("gene_count_all", ""),
            v.get("gene_count_up", ""),
            v.get("gene_count_down", ""),
            v.get("gene_ids_all", ""),
            v.get("gene_ids_up", ""),
            v.get("gene_ids_down", ""),
            v.get("bg_size", ""),
            v.get("universe_size", ""),
            v.get("min_gs", ""),
            v.get("max_gs", ""),
            v.get("count_mode", ""),
        ]
        ws.append(row)

    logging.info("contrast=%s：已在 xlsx 中添加 KEGG_summary 宽表（%d 个 pathway）。", contrast, len(summary))


def build_contrast_xlsx(
    contrast: str,
    source_data_root: Path,
    publish_root: Path,
    manifest_rows: List[Dict[str, str]],
) -> None:
    """
    为单个 contrast 生成一个 <contrast>.xlsx 文件。

    只依赖 source_data 下的内容：
      - deg/<contrast>/DEG_all.tsv
      - enrich/<contrast>/GO_sig_up.tsv
      - enrich/<contrast>/GO_sig_down.tsv
      - enrich/<contrast>/KEGG_by_term_all.tsv
      - background/<contrast>.meta.tsv

    以及宽表：
      - enrich/<contrast>/GO_*_by_term_*.tsv  -> GO_summary
      - enrich/<contrast>/KEGG_by_term_*.tsv  -> KEGG_summary

    任意文件缺失则跳过对应 sheet，不报错。
    """
    try:
        from openpyxl import Workbook  # 延迟导入，避免环境没有 openpyxl 时崩溃
    except ImportError:
        logging.warning("未安装 openpyxl，无法生成 xlsx（contrast=%s）。", contrast)
        return

    deg_all_tsv = source_data_root / "deg" / contrast / "DEG_all.tsv"
    go_sig_up_tsv = source_data_root / "enrich" / contrast / "GO_sig_up.tsv"
    go_sig_down_tsv = source_data_root / "enrich" / contrast / "GO_sig_down.tsv"
    kegg_all_tsv = source_data_root / "enrich" / contrast / "KEGG_by_term_all.tsv"
    bg_meta_tsv = source_data_root / "background" / f"{contrast}.meta.tsv"
    design_txt = source_data_root / "deg" / contrast / "design.txt"
    go_sig_all_tsv = source_data_root / "enrich" / contrast / "GO_sig_all.tsv"

    available = any([
        deg_all_tsv.is_file(),
        go_sig_up_tsv.is_file(),
        go_sig_down_tsv.is_file(),
        kegg_all_tsv.is_file(),
        bg_meta_tsv.is_file(),
        design_txt.is_file(),
        go_sig_all_tsv.is_file(),
    ])
    if not available:
        logging.info("contrast=%s 没有可用的 TSV 供生成 xlsx，跳过。", contrast)
        return

    wb = Workbook()
    default_ws = wb.active
    default_ws.title = "DEG_all"

    # DEG_all sheet
    if deg_all_tsv.is_file():
        ws = wb["DEG_all"]
        tsv_to_worksheet(ws, deg_all_tsv)
    else:
        wb.remove(default_ws)

    # GO_sig_up sheet
    if go_sig_up_tsv.is_file():
        ws = wb.create_sheet(title="GO_sig_up")
        tsv_to_worksheet(ws, go_sig_up_tsv)

    # GO_sig_down sheet
    if go_sig_down_tsv.is_file():
        ws = wb.create_sheet(title="GO_sig_down")
        tsv_to_worksheet(ws, go_sig_down_tsv)

    # GO_sig_all sheet
    if go_sig_all_tsv.is_file():
        # 即使只有表头也照样写入，方便后续确认口径
        ws = wb.create_sheet(title="GO_sig_all")
        tsv_to_worksheet(ws, go_sig_all_tsv)

    # KEGG_all sheet（用 KEGG_by_term_all.tsv）
    if kegg_all_tsv.is_file():
        ws = wb.create_sheet(title="KEGG_all")
        tsv_to_worksheet(ws, kegg_all_tsv)

    # background_meta sheet
    if bg_meta_tsv.is_file():
        ws = wb.create_sheet(title="background_meta")
        tsv_to_worksheet(ws, bg_meta_tsv)

    # design sheet
    if design_txt.is_file():
        ws = wb.create_sheet(title="design")
        with design_txt.open("r", encoding="utf-8") as f:
            for row_idx, line in enumerate(f, start=1):
                ws.cell(row=row_idx, column=1, value=line.rstrip("\n"))

    # GO_summary 宽表
    build_go_summary_sheet(wb, contrast, source_data_root)

    # KEGG_summary 宽表
    build_kegg_summary_sheet(wb, contrast, source_data_root)

    # 若最终 workbook 中没有任何 sheet，就不写文件
    if len(wb.worksheets) == 0:
        logging.info("contrast=%s workbook 为空，未生成 xlsx。", contrast)
        return

    xlsx_path = publish_root / f"{contrast}.xlsx"
    wb.save(xlsx_path)
    logging.info("已生成 xlsx：%s", xlsx_path)

    # 记录到 manifest（这里用相对于 publish_root 的路径看起来更直观）
    try:
        rel = xlsx_path.relative_to(source_data_root)
    except ValueError:
        # xlsx 不在 source_data 下，用相对于 publish_root 的路径记录
        try:
            rel = xlsx_path.relative_to(publish_root)
        except ValueError:
            rel = xlsx_path

    manifest_rows.append(
        {
            "path": str(rel),
            "module": "xlsx",
            "label": contrast,
            "description": "Excel workbook summarising key tables and wide GO/KEGG summaries for this contrast",
        }
    )


def main() -> None:
    # ---------- 初始化 ----------
    setup_logging("INFO")
    cfg = load_config(Path(CONFIG_FILE))

    dirs_cfg = cfg.get("dirs", {})
    publish_cfg = cfg.get("publish", {})

    matrix_dir = Path(dirs_cfg.get("matrix", "results/05_matrix"))
    deg_dir = Path(dirs_cfg.get("deg", "results/06_deg"))
    enrich_dir = Path(dirs_cfg.get("enrich", "results/08_enrich"))

    publish_root = Path(dirs_cfg.get("publish", "results/09_publish"))
    source_data_root = Path(publish_cfg.get("source_data_dir", str(publish_root / "source_data")))
    methods_readme_path = Path(publish_cfg.get("methods_readme", str(source_data_root / "METHODS_README.txt")))
    enable_xlsx = bool(publish_cfg.get("xlsx", False))

    ensure_dir(publish_root)
    ensure_dir(source_data_root)

    logging.info("发布目录（publish_root）：%s", publish_root)
    logging.info("Source Data 根目录：%s", source_data_root)
    logging.info("publish.xlsx 开关：%s", "true" if enable_xlsx else "false")

    manifest_rows: List[Dict[str, str]] = []
    check_rows: List[Dict[str, str]] = []

# ========== 1) matrix 模块 ==========
    logging.info("===== 1) 打包 matrix 模块 =====")
    src_counts = matrix_dir / "counts" / "gene_counts.tsv"
    src_tpm = matrix_dir / "tpms" / "gene_tpm.tsv"
    src_stats = matrix_dir / "matrix_stats.tsv"

    matrix_out_dir = source_data_root / "matrix"
    ensure_dir(matrix_out_dir)

    if src_counts.is_file():
        dst = matrix_out_dir / "gene_counts.tsv"
        copy_file(
            src_counts,
            dst,
            module="matrix",
            label="",
            description="Gene-level counts matrix (DESeq2 input)",
            manifest_rows=manifest_rows,
            source_data_root=source_data_root,
        )
    else:
        check_rows.append(
            {"module": "matrix", "label": "NA", "status": "missing", "detail": f"missing file: {src_counts}"}
        )
        logging.warning("缺少 matrix 文件：%s", src_counts)

    if src_tpm.is_file():
        dst = matrix_out_dir / "gene_tpm.tsv"
        copy_file(
            src_tpm,
            dst,
            module="matrix",
            label="",
            description="Gene-level TPM matrix",
            manifest_rows=manifest_rows,
            source_data_root=source_data_root,
        )
    else:
        check_rows.append(
            {"module": "matrix", "label": "NA", "status": "missing", "detail": f"missing file: {src_tpm}"}
        )
        logging.warning("缺少 matrix 文件：%s", src_tpm)

    if src_stats.is_file():
        dst = matrix_out_dir / "matrix_stats.tsv"
        copy_file(
            src_stats,
            dst,
            module="matrix",
            label="",
            description="Matrix-level statistics (number of genes, library sizes, etc.)",
            manifest_rows=manifest_rows,
            source_data_root=source_data_root,
        )
    else:
        check_rows.append(
            {"module": "matrix", "label": "NA", "status": "missing", "detail": f"missing file: {src_stats}"}
        )
        logging.warning("缺少 matrix 文件：%s", src_stats)

    # ========== 2) DEG 模块 ==========
    logging.info("===== 2) 打包 DEG 模块 =====")
    contrasts = get_contrasts_list(cfg)
    if not contrasts:
        logging.warning("没有从 contrasts.tsv 中获取到对比列表，DEG 模块将不会打包具体对比。")

    deg_out_root = source_data_root / "deg"
    ensure_dir(deg_out_root)

    for contrast in contrasts:
        contrast_dir = deg_dir / contrast
        if not contrast_dir.is_dir():
            check_rows.append(
                {
                    "module": "deg",
                    "label": contrast,
                    "status": "missing",
                    "detail": f"DEG directory not found: {contrast_dir}",
                }
            )
            logging.warning("DEG 目录缺失：%s", contrast_dir)
            continue

        deg_out_dir = deg_out_root / contrast
        ensure_dir(deg_out_dir)

        expected_files = [
            ("design.txt", "DESeq2 design formula and covariates"),
            ("DEG_all.tsv", "All DEGs for this contrast"),
            ("DEG_up.tsv", "Up-regulated DEGs for this contrast"),
            ("DEG_down.tsv", "Down-regulated DEGs for this contrast"),
            ("varTop100.list", "Top variable genes for diagnostics"),
            ("rle_range.tsv", "RLE summary for normalisation diagnostics"),
        ]

        missing_list: List[str] = []

        for fname, desc in expected_files:
            src = contrast_dir / fname
            if src.is_file():
                dst = deg_out_dir / fname
                copy_file(
                    src,
                    dst,
                    module="deg",
                    label=contrast,
                    description=desc,
                    manifest_rows=manifest_rows,
                    source_data_root=source_data_root,
                )
            else:
                missing_list.append(str(src))
                logging.warning("DEG 文件缺失：%s", src)

        if missing_list:
            check_rows.append(
                {
                    "module": "deg",
                    "label": contrast,
                    "status": "partial",
                    "detail": "missing files: " + ", ".join(missing_list),
                }
            )
        else:
            check_rows.append({"module": "deg", "label": contrast, "status": "ok", "detail": ""})

    # ========== 3) 背景模块 ==========
    logging.info("===== 3) 打包背景模块 =====")
    background_in_dir = enrich_dir / "background"
    background_out_dir = source_data_root / "background"
    ensure_dir(background_out_dir)

    for contrast in contrasts:
        list_in = background_in_dir / f"{contrast}.list"
        meta_in = background_in_dir / f"{contrast}.meta.tsv"
        missing_bg: List[str] = []

        if list_in.is_file():
            list_out = background_out_dir / f"{contrast}.list"
            copy_file(
                list_in,
                list_out,
                module="background",
                label=contrast,
                description="Background universe gene_id list for this contrast",
                manifest_rows=manifest_rows,
                source_data_root=source_data_root,
            )
        else:
            missing_bg.append(str(list_in))
            logging.warning("背景 list 缺失：%s", list_in)

        if meta_in.is_file():
            meta_out = background_out_dir / f"{contrast}.meta.tsv"
            copy_file(
                meta_in,
                meta_out,
                module="background",
                label=contrast,
                description="Background metadata (universe_size, coverage, sample subset)",
                manifest_rows=manifest_rows,
                source_data_root=source_data_root,
            )
        else:
            missing_bg.append(str(meta_in))
            logging.warning("背景 meta 缺失：%s", meta_in)

        if missing_bg:
            check_rows.append(
                {
                    "module": "background",
                    "label": contrast,
                    "status": "partial",
                    "detail": "missing files: " + ", ".join(missing_bg),
                }
            )
        else:
            check_rows.append({"module": "background", "label": contrast, "status": "ok", "detail": ""})

    # ========== 4) 富集模块 ==========
    logging.info("===== 4) 打包富集模块 =====")
    enrich_out_root = source_data_root / "enrich"
    ensure_dir(enrich_out_root)

    enrich_labels: List[str] = []
    if enrich_dir.is_dir():
        for child in enrich_dir.iterdir():
            if not child.is_dir():
                continue
            name = child.name
            if name in {"inputs", "background"}:
                continue
            enrich_labels.append(name)
    enrich_labels = sorted(set(enrich_labels))
    logging.info("检测到富集 label 数量：%d（%s）", len(enrich_labels), ", ".join(enrich_labels) if enrich_labels else "无")

    onts = ["BP", "CC", "MF"]
    sets = ["all", "up", "down"]

    for label in enrich_labels:
        label_in_dir = enrich_dir / label
        label_out_dir = enrich_out_root / label
        ensure_dir(label_out_dir)

        missing_enrich: List[str] = []

        # GO by-term
        for ont in onts:
            for test_set in sets:
                fname = f"GO_{ont}_by_term_{test_set}.tsv"
                src = label_in_dir / fname
                if src.is_file():
                    dst = label_out_dir / fname
                    desc = f"GO {ont} enrichment by-term table ({test_set}) for label {label}"
                    copy_file(
                        src,
                        dst,
                        module="enrich",
                        label=label,
                        description=desc,
                        manifest_rows=manifest_rows,
                        source_data_root=source_data_root,
                    )
                else:
                    missing_enrich.append(str(src))

        # GO_sig_*
        for set_name in ["all", "up", "down"]:
            fname = f"GO_sig_{set_name}.tsv"
            src = label_in_dir / fname
            if src.is_file():
                dst = label_out_dir / fname
                desc = f"GO significant terms ({set_name}) for label {label}"
                copy_file(
                    src,
                    dst,
                    module="enrich",
                    label=label,
                    description=desc,
                    manifest_rows=manifest_rows,
                    source_data_root=source_data_root,
                )
            else:
                missing_enrich.append(str(src))

        # KEGG by-term
        for test_set in sets:
            fname = f"KEGG_by_term_{test_set}.tsv"
            src = label_in_dir / fname
            if src.is_file():
                dst = label_out_dir / fname
                desc = f"KEGG enrichment by-term table ({test_set}) for label {label}"
                copy_file(
                    src,
                    dst,
                    module="enrich",
                    label=label,
                    description=desc,
                    manifest_rows=manifest_rows,
                    source_data_root=source_data_root,
                )
            else:
                missing_enrich.append(str(src))

        # GSEA 可选
        gsea_in_dir = label_in_dir / "gsea"
        if gsea_in_dir.is_dir():
            gsea_out_dir = label_out_dir / "gsea"
            ensure_dir(gsea_out_dir)
            for gsea_fname, desc in [
                ("GO_gsea.tsv", f"GO GSEA results for label {label}"),
                ("KEGG_gsea.tsv", f"KEGG GSEA results for label {label}"),
            ]:
                src = gsea_in_dir / gsea_fname
                if src.is_file():
                    dst = gsea_out_dir / gsea_fname
                    copy_file(
                        src,
                        dst,
                        module="enrich",
                        label=label,
                        description=desc,
                        manifest_rows=manifest_rows,
                        source_data_root=source_data_root,
                    )
                else:
                    missing_enrich.append(str(src))

        if missing_enrich:
            check_rows.append(
                {
                    "module": "enrich",
                    "label": label,
                    "status": "partial_or_missing",
                    "detail": "missing (or not generated) files: " + ", ".join(missing_enrich),
                }
            )
        else:
            check_rows.append({"module": "enrich", "label": label, "status": "ok", "detail": ""})

    # ========== 5) manifest.tsv & publish_check.tsv ==========
    logging.info("===== 5) 写出 manifest.tsv 与 publish_check.tsv =====")
    manifest_path = source_data_root / "manifest.tsv"
    write_tsv(manifest_path, manifest_rows, MANIFEST_COLUMNS)

    check_path = source_data_root / "publish_check.tsv"
    write_tsv(check_path, check_rows, CHECK_COLUMNS)

    # ========== 6) METHODS_README.txt ==========
    logging.info("===== 6) 生成 METHODS_README.txt =====")
    samples_summary = summarize_samples(cfg)
    methods_text = build_methods_readme_text(cfg, samples_summary, source_data_root)

    ensure_dir(methods_readme_path.parent)
    with methods_readme_path.open("w", encoding="utf-8") as f:
        f.write(methods_text)

    # 把 METHODS_README 也加进 manifest
    try:
        rel_path = methods_readme_path.relative_to(source_data_root)
    except ValueError:
        rel_path = methods_readme_path
    manifest_rows.append(
        {
            "path": str(rel_path),
            "module": "methods",
            "label": "NA",
            "description": "Methods README skeleton for this RNA-seq project",
        }
    )
    write_tsv(manifest_path, manifest_rows, MANIFEST_COLUMNS)

    # ========== 7) 可选：生成每个 contrast 的 xlsx ==========
    if enable_xlsx and contrasts:
        logging.info("===== 7) 生成 per-contrast xlsx 工作簿（含 GO/KEGG 宽表） =====")
        for contrast in contrasts:
            build_contrast_xlsx(
                contrast=contrast,
                source_data_root=source_data_root,
                publish_root=publish_root,
                manifest_rows=manifest_rows,
            )
        # xlsx 也写入 manifest 后，再次刷新 manifest.tsv
        write_tsv(manifest_path, manifest_rows, MANIFEST_COLUMNS)
    else:
        if not enable_xlsx:
            logging.info("publish.xlsx = false，不生成 xlsx。")
        elif not contrasts:
            logging.info("未检测到对比列表，跳过 xlsx 生成。")

    logging.info("===== 09_publish_results.py 执行完成，发布结果已写入：%s =====", source_data_root)


if __name__ == "__main__":
    main()
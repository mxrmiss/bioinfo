#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
07b_make_psg_qsig_bebsig_table.py

用途：
- 不修改原 07 脚本与其输出；
- 读取原 07 产物：
    results/05_cmlagg/D_fdr_genes.tsv
    results/05_cmlagg/D_beb_sites.tsv
- 过滤出：q 显著（Q <= fdr_alpha）且存在 BEB 位点显著（post >= cutoff）的 OG×FG
- 额外从 codeml raw 结果补齐：lnL0(null)、lnL1(alt)、kappa
- 生成最终论文风格表：
    PSG_qsig_bebsig_OGxFG.tsv
    PSG_qsig_bebsig_OGxFG.xlsx

列顺序（皇上定稿）：
  OG, GeneNames(=纯ID串), foreground, foreground_gene_ids, lnL0, lnL1, kappa, LRT, pvalue, qvalue,
  PositiveSites, PosteriorDetail
"""

from __future__ import annotations
import sys
import re
import logging
from pathlib import Path
from typing import Dict, Any, Tuple, List
import yaml
from collections import defaultdict

# ====== 新增：写 Excel（openpyxl） ======
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ===================== 皇上在这里填参数（不走命令行） =====================

CONFIG_PATH = "config.yaml"  # 默认读取项目根目录的 config.yaml

OUTPUT_TSV = "PSG_qsig_bebsig_OGxFG.tsv"
OUTPUT_XLSX = "PSG_qsig_bebsig_OGxFG.xlsx"
WRITE_XLSX = True  # 想只要 TSV 就改成 False

# “位点显著”阈值（与你 07 中 BEB cutoff 保持一致）
BEB_POST_CUTOFF = 0.95

# q 显著阈值：若 config 里能读到就用 config 的，否则用这里默认
FDR_ALPHA_FALLBACK = 0.05

# 输出位点/PosteriorDetail 的小数位数
POST_DECIMALS = 3

# ===================== 通用工具 =====================

FLOAT_RE = re.compile(r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?")

def read_yaml(fp: Path) -> Dict[str, Any]:
    with fp.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}

def ensure_dir(p: Path) -> Path:
    p.mkdir(parents=True, exist_ok=True)
    return p

def need_file(p: Path, msg: str) -> Path:
    if not p.is_file():
        raise FileNotFoundError(f"[ERR] 缺少 {msg}: {p}")
    return p

def need_dir(p: Path, msg: str) -> Path:
    if not p.is_dir():
        raise FileNotFoundError(f"[ERR] 缺少 {msg}: {p}")
    return p

def get_logger(log_file: Path) -> logging.Logger:
    ensure_dir(log_file.parent)
    log = logging.getLogger("aphylo.07b")
    log.setLevel(logging.INFO)
    if not log.handlers:
        fmt = logging.Formatter("[%(asctime)s] [%(levelname)s] %(message)s")
        sh = logging.StreamHandler(sys.stdout)
        sh.setFormatter(fmt)
        fh = logging.FileHandler(log_file, encoding="utf-8")
        fh.setFormatter(fmt)
        log.addHandler(sh)
        log.addHandler(fh)
    return log

def last_lnl(txt: str) -> float | None:
    """
    解析 mlc 文本最后一次出现的 lnL。
    兼容：
      lnL(ntime: ...): -123.456
      lnL = -123.456
    """
    NUM = r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?"
    m1 = re.findall(rf"^lnL\s*\([^\n]*?\)\s*:\s*({NUM})", txt, flags=re.M)
    m2 = re.findall(rf"^lnL\s*=\s*({NUM})", txt, flags=re.M)
    m = (m1 if m1 else []) + (m2 if m2 else [])
    if not m:
        return None
    try:
        return float(m[-1])
    except Exception:
        return None

def parse_kappa(txt: str) -> float | None:
    """
    解析 kappa（ts/tv）。
    常见：
      kappa (ts/tv) =  1.234
      kappa = 1.234
    """
    NUM = r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?"
    pats = [
        rf"^kappa\s*\(ts\/tv\)\s*=\s*({NUM})",
        rf"^kappa\s*=\s*({NUM})",
    ]
    hits: List[str] = []
    for p in pats:
        hits += re.findall(p, txt, flags=re.M)
    if not hits:
        return None
    try:
        return float(hits[-1])
    except Exception:
        return None

def get_fdr_alpha(cfg: Dict[str, Any]) -> float:
    """
    阈值优先级：
      1) config.psg.fdr_alpha
      2) config.report.fdr_alpha
      3) FDR_ALPHA_FALLBACK
    """
    psg_cfg = (cfg.get("psg") or {})
    report_cfg = (cfg.get("report") or {})
    val = psg_cfg.get("fdr_alpha", None)
    if val is None:
        val = report_cfg.get("fdr_alpha", None)
    if val is None:
        return float(FDR_ALPHA_FALLBACK)
    try:
        return float(val)
    except Exception:
        return float(FDR_ALPHA_FALLBACK)

def read_tsv(fp: Path) -> List[List[str]]:
    rows: List[List[str]] = []
    with fp.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue
            rows.append(line.split("\t"))
    return rows

def write_xlsx(out_fp: Path, header: List[str], rows: List[List[str]], sheet_name: str = "PSG"):
    """
    写 Excel：第一行为表头，后面逐行写入。
    注意：全部以“字符串”写入，避免 WPS/Excel 自动把 OG 变科学计数法。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet 名最长 31

    # 写表头
    ws.append([str(x) for x in header])

    # 写数据
    for r in rows:
        ws.append([str(x) if x is not None else "" for x in r])

    # 简单设置列宽（按最长内容截断到上限）
    max_width = 60
    for col_idx in range(1, len(header) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header[col_idx - 1]))
        for row in rows[:2000]:  # 只抽样前 2000 行估算，避免太慢
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[col_letter].width = min(max_width, max(10, max_len + 2))

    ensure_dir(out_fp.parent)
    wb.save(out_fp)

# ===================== 主流程 =====================

def main():
    cfg_fp = need_file(Path(CONFIG_PATH), "config.yaml")
    cfg = read_yaml(cfg_fp)

    paths = cfg.get("paths", {}) or {}
    if not paths:
        raise KeyError("[ERR] config.yaml 缺少 paths 节。")

    logs_dir = Path(paths["logs_dir"])
    log = get_logger(logs_dir / "07b_make_psg_qsig_bebsig_table.log")

    codeml_agg_dir = need_dir(Path(paths["codeml_agg_dir"]), "codeml_agg_dir")
    codeml_dir = need_dir(Path(paths["codeml_dir"]), "codeml_dir")
    raw_root = need_dir(codeml_dir / "raw", "codeml raw 目录")

    fdr_alpha = get_fdr_alpha(cfg)
    log.info("=" * 80)
    log.info("APhylo 07b — 生成 PSG_qsig_bebsig_OGxFG（TSV + XLSX，可选）")
    log.info(f"[PARAM] fdr_alpha={fdr_alpha}  BEB_POST_CUTOFF={BEB_POST_CUTOFF}  WRITE_XLSX={WRITE_XLSX}")
    log.info("=" * 80)

    fp_genes = need_file(codeml_agg_dir / "D_fdr_genes.tsv", "D_fdr_genes.tsv（来自原 07）")
    fp_sites = need_file(codeml_agg_dir / "D_beb_sites.tsv", "D_beb_sites.tsv（来自原 07）")

    # ---------- 读取基因层表 ----------
    gene_rows = read_tsv(fp_genes)
    if not gene_rows:
        raise ValueError(f"[ERR] 空文件：{fp_genes}")

    header_g = gene_rows[0]
    idx_g = {name: i for i, name in enumerate(header_g)}

    need_cols_g = ["OG", "foreground", "LRT", "P", "Q"]
    for c in need_cols_g:
        if c not in idx_g:
            raise KeyError(f"[ERR] D_fdr_genes.tsv 缺少列：{c}")

    has_fgids = "foreground_gene_ids" in idx_g

    # (OG, FG) -> {LRT,P,Q,fgids}
    gene_map: Dict[Tuple[str, str], Dict[str, str]] = {}
    for r in gene_rows[1:]:
        if len(r) < len(header_g):
            continue
        og = r[idx_g["OG"]]
        fg = r[idx_g["foreground"]]
        lrt_s = r[idx_g["LRT"]]
        p_s = r[idx_g["P"]]
        q_s = r[idx_g["Q"]]
        fgids = r[idx_g["foreground_gene_ids"]] if has_fgids else ""
        gene_map[(og, fg)] = {
            "LRT": lrt_s,
            "P": p_s,
            "Q": q_s,
            "foreground_gene_ids": fgids,
        }

    # ---------- 读取位点表 ----------
    site_rows = read_tsv(fp_sites)
    if not site_rows:
        raise ValueError(f"[ERR] 空文件：{fp_sites}")

    header_s = site_rows[0]
    idx_s = {name: i for i, name in enumerate(header_s)}
    need_cols_s = ["OG", "foreground", "site", "aa", "post"]
    for c in need_cols_s:
        if c not in idx_s:
            raise KeyError(f"[ERR] D_beb_sites.tsv 缺少列：{c}")

    has_fgids_s = "foreground_gene_ids" in idx_s

    # (OG, FG) -> list of (site, aa, post) 仅保留 post>=cutoff
    beb_map: Dict[Tuple[str, str], List[Tuple[int, str, float]]] = defaultdict(list)
    for r in site_rows[1:]:
        if len(r) < len(header_s):
            continue
        og = r[idx_s["OG"]]
        fg = r[idx_s["foreground"]]
        try:
            site_i = int(r[idx_s["site"]])
        except Exception:
            continue
        aa = r[idx_s["aa"]]
        try:
            post_f = float(r[idx_s["post"]])
        except Exception:
            continue
        if post_f >= float(BEB_POST_CUTOFF):
            beb_map[(og, fg)].append((site_i, aa, post_f))

        # 若 gene_map 没有 fgids，但 sites 有，拿 sites 的补一下
        if (og, fg) in gene_map and (not gene_map[(og, fg)].get("foreground_gene_ids")) and has_fgids_s:
            gene_map[(og, fg)]["foreground_gene_ids"] = r[idx_s["foreground_gene_ids"]]

    # ---------- 生成最终表 ----------
    out_tsv = codeml_agg_dir / OUTPUT_TSV
    out_xlsx = codeml_agg_dir / OUTPUT_XLSX

    header_final = [
        "OG", "GeneNames", "foreground", "foreground_gene_ids", "lnL0", "lnL1", "kappa",
        "LRT", "pvalue", "qvalue", "PositiveSites", "PosteriorDetail"
    ]

    lines_tsv: List[str] = []
    rows_xlsx: List[List[str]] = []

    kept = 0
    for (og, fg), g in gene_map.items():
        # q 显著
        try:
            qv = float(g["Q"])
        except Exception:
            continue
        if qv > fdr_alpha:
            continue

        # 位点显著（至少一个）
        sites = beb_map.get((og, fg), [])
        if not sites:
            continue

        # 解析 raw 的 lnL/kappa
        alt_mlc = need_file(raw_root / og / fg / "alt" / "mlc.txt", f"ALT mlc ({og}/{fg})")
        null_mlc = need_file(raw_root / og / fg / "null" / "mlc.txt", f"NULL mlc ({og}/{fg})")
        alt_txt = alt_mlc.read_text(encoding="utf-8", errors="ignore")
        nul_txt = null_mlc.read_text(encoding="utf-8", errors="ignore")

        lnL1 = last_lnl(alt_txt)
        lnL0 = last_lnl(nul_txt)
        if lnL1 is None or lnL0 is None:
            raise ValueError(f"[ERR] 无法解析 lnL：{alt_mlc} 或 {null_mlc}")

        kappa = parse_kappa(alt_txt)
        if kappa is None:
            kappa = parse_kappa(nul_txt)

        # 排序并拼接位点信息
        sites.sort(key=lambda x: x[0])
        pos_sites = ",".join(str(s[0]) for s in sites) if sites else "None"
        fmt_post = f"{{:.{POST_DECIMALS}f}}"
        posterior_detail = ",".join(f"{s[0]}({s[1]}):{fmt_post.format(s[2])}" for s in sites)

        fgids = g.get("foreground_gene_ids", "")
        gene_names = fgids  # 皇上定稿：GeneNames 直接用纯 ID 串，不带 (ATP2A)

        row = [
            og,
            gene_names,
            fg,
            fgids,
            f"{lnL0:.6f}",
            f"{lnL1:.6f}",
            "" if kappa is None else f"{kappa:.5f}",
            g["LRT"],
            g["P"],
            f"{qv:.6g}",
            pos_sites,
            posterior_detail,
        ]

        rows_xlsx.append(row)
        lines_tsv.append("\t".join(row))
        kept += 1

    # 写 TSV
    out_tsv.write_text("\t".join(header_final) + "\n" + ("\n".join(lines_tsv) + "\n" if lines_tsv else ""), encoding="utf-8")
    log.info(f"[OK] 写出 TSV：{out_tsv}（{kept} 条）")

    # 写 XLSX
    if WRITE_XLSX:
        write_xlsx(out_xlsx, header_final, rows_xlsx, sheet_name="PSG_qsig_bebsig")
        log.info(f"[OK] 写出 XLSX：{out_xlsx}（{kept} 条）")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        sys.stderr.write(str(e) + "\n")
        sys.exit(2)

